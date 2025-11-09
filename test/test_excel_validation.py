"""
Script de prueba para generar un Excel de estados financieros con valores aleatorios
y validar que la carga funcione correctamente.
"""
import os
import sys
import django
import random
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Protection, Border, Side
from openpyxl.utils import get_column_letter

# Configurar Django
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'sistema_financiero.settings')
django.setup()

from stela.models.catalogo import Catalogo, GrupoCuenta, Cuenta
from stela.models.empresa import Empresa
from stela.models.finanzas import Periodo, Balance, BalanceDetalle
from stela.services.plantillas import generar_plantilla_estados_excel


def generar_valores_aleatorios():
    """Genera valores aleatorios para Debe y Haber"""
    # Generar valores entre 0 y 1,000,000
    debe = Decimal(str(random.randint(0, 1000000)))
    haber = Decimal(str(random.randint(0, 1000000)))
    return debe, haber


def crear_excel_prueba(catalogo, output_path="test_estados_financieros.xlsx"):
    """
    Crea un Excel de prueba basado en la plantilla generada,
    pero con valores aleatorios en Debe y Haber.
    """
    print(f"Generando Excel de prueba para catálogo: {catalogo}")
    
    # Obtener cuentas del catálogo
    grupos = GrupoCuenta.objects.filter(catalogo=catalogo).select_related()
    
    # Agrupar cuentas por bg_bloque para Balance General
    cuentas_por_bg_bloque = {
        'ACTIVO_CORRIENTE': [],
        'ACTIVO_NO_CORRIENTE': [],
        'PASIVO_CORRIENTE': [],
        'PASIVO_NO_CORRIENTE': [],
        'PATRIMONIO': [],
    }
    
    # Agrupar cuentas por er_bloque para Estado de Resultados
    cuentas_por_er_bloque = {
        'VENTAS_NETAS': [],
        'COSTO_NETO_VENTAS': [],
        'GASTOS_OPERATIVOS': [],
        'OTROS_INGRESOS': [],
        'OTROS_GASTOS': [],
        'GASTO_FINANCIERO': [],
        'IMPUESTO_SOBRE_LA_RENTA': [],
    }
    
    for grupo in grupos:
        cuentas = Cuenta.objects.filter(grupo=grupo)
        for cuenta in cuentas:
            if cuenta.bg_bloque and cuenta.bg_bloque in cuentas_por_bg_bloque:
                cuentas_por_bg_bloque[cuenta.bg_bloque].append(cuenta)
            if cuenta.er_bloque and cuenta.er_bloque in cuentas_por_er_bloque:
                cuentas_por_er_bloque[cuenta.er_bloque].append(cuenta)
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True)
    debe_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    haber_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    subtotal_border = Border(top=Side(style="thin"))
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # ========== HOJA 1: BALANCE GENERAL ==========
    ws_balance = wb.create_sheet("BalanceGeneral")
    
    # Leyenda
    ws_balance.append(['LEYENDA:', 'Verde claro = Debe (principal para Activos)', 'Azul claro = Haber (principal para Pasivos/Patrimonio)', '', '', '', '', ''])
    ws_balance.merge_cells('A1:D1')
    ws_balance.cell(row=1, column=1).font = Font(bold=True, size=10)
    ws_balance.cell(row=1, column=1).fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    ws_balance.cell(row=1, column=2).fill = debe_fill
    ws_balance.cell(row=1, column=3).fill = haber_fill
    
    # Encabezados
    ws_balance.append(['Activos', '', '', '', 'Pasivos y Patrimonio', '', '', ''])
    ws_balance.append(['Código', 'Nombre', 'Debe', 'Haber', 'Código', 'Nombre', 'Debe', 'Haber'])
    
    for row in [2, 3]:
        for col in range(1, 9):
            cell = ws_balance.cell(row=row, column=col)
            if row == 2:
                cell.fill = header_fill
                cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
    
    nombres_bg_bloques = {
        'ACTIVO_CORRIENTE': 'Activo Corriente',
        'ACTIVO_NO_CORRIENTE': 'Activo No Corriente',
        'PASIVO_CORRIENTE': 'Pasivo Corriente',
        'PASIVO_NO_CORRIENTE': 'Pasivo No Corriente',
        'PATRIMONIO': 'Patrimonio'
    }
    
    row_num_activos = 4
    row_num_pasivos = 4
    
    # Activos (columna izquierda)
    for bloque_key in ['ACTIVO_CORRIENTE', 'ACTIVO_NO_CORRIENTE']:
        if bloque_key in cuentas_por_bg_bloque and cuentas_por_bg_bloque[bloque_key]:
            nombre_bloque = nombres_bg_bloques.get(bloque_key, bloque_key)
            ws_balance.cell(row=row_num_activos, column=1, value=nombre_bloque).font = bold_font
            row_num_activos += 1
            
            for cuenta in cuentas_por_bg_bloque[bloque_key]:
                debe, haber = generar_valores_aleatorios()
                ws_balance.cell(row=row_num_activos, column=1, value=cuenta.codigo)
                ws_balance.cell(row=row_num_activos, column=2, value=cuenta.nombre)
                ws_balance.cell(row=row_num_activos, column=3, value=float(debe)).fill = debe_fill
                ws_balance.cell(row=row_num_activos, column=4, value=float(haber)).fill = haber_fill
                row_num_activos += 1
    
    # Pasivos y Patrimonio (columna derecha)
    for bloque_key in ['PASIVO_CORRIENTE', 'PASIVO_NO_CORRIENTE', 'PATRIMONIO']:
        if bloque_key in cuentas_por_bg_bloque and cuentas_por_bg_bloque[bloque_key]:
            nombre_bloque = nombres_bg_bloques.get(bloque_key, bloque_key)
            ws_balance.cell(row=row_num_pasivos, column=5, value=nombre_bloque).font = bold_font
            row_num_pasivos += 1
            
            for cuenta in cuentas_por_bg_bloque[bloque_key]:
                debe, haber = generar_valores_aleatorios()
                ws_balance.cell(row=row_num_pasivos, column=5, value=cuenta.codigo)
                ws_balance.cell(row=row_num_pasivos, column=6, value=cuenta.nombre)
                ws_balance.cell(row=row_num_pasivos, column=7, value=float(debe)).fill = debe_fill
                ws_balance.cell(row=row_num_pasivos, column=8, value=float(haber)).fill = haber_fill
                row_num_pasivos += 1
    
    # Ajustar ancho de columnas
    for col in range(1, 9):
        ws_balance.column_dimensions[get_column_letter(col)].width = 20
    
    # ========== HOJA 2: ESTADO DE RESULTADOS ==========
    ws_resultados = wb.create_sheet("EstadoResultados")
    
    # Leyenda
    ws_resultados.append(['LEYENDA:', 'Verde claro = Debe (principal para Gastos)', 'Azul claro = Haber (principal para Ingresos)', ''])
    ws_resultados.merge_cells('A1:D1')
    ws_resultados.cell(row=1, column=1).font = Font(bold=True, size=10)
    ws_resultados.cell(row=1, column=1).fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    ws_resultados.cell(row=1, column=2).fill = debe_fill
    ws_resultados.cell(row=1, column=3).fill = haber_fill
    
    # Encabezados
    ws_resultados.append(['Código', 'Cuenta', 'Debe', 'Haber', 'Total'])
    
    for col in range(1, 6):
        cell = ws_resultados.cell(row=2, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    orden_bloques_er = [
        'VENTAS_NETAS',
        'COSTO_NETO_VENTAS',
        'GASTOS_OPERATIVOS',
        'OTROS_INGRESOS',
        'OTROS_GASTOS',
        'GASTO_FINANCIERO',
        'IMPUESTO_SOBRE_LA_RENTA'
    ]
    
    nombres_bloques_er = {
        'VENTAS_NETAS': 'Ventas Netas',
        'COSTO_NETO_VENTAS': 'Costo Neto de Ventas',
        'GASTOS_OPERATIVOS': 'Gastos Operativos',
        'OTROS_INGRESOS': 'Otros Ingresos',
        'OTROS_GASTOS': 'Otros Gastos',
        'GASTO_FINANCIERO': 'Gasto Financiero',
        'IMPUESTO_SOBRE_LA_RENTA': 'Impuesto sobre la Renta'
    }
    
    row_num = 3
    
    # Agregar cuentas por bloque ER
    for bloque_key in orden_bloques_er:
        cuentas = cuentas_por_er_bloque.get(bloque_key, [])
        if not cuentas:
            continue
        
        for cuenta in cuentas:
            debe, haber = generar_valores_aleatorios()
            # Código (columna 1)
            ws_resultados.cell(row=row_num, column=1, value=cuenta.codigo)
            # Nombre indentado (columna 2)
            name_cell = ws_resultados.cell(row=row_num, column=2, value=cuenta.nombre)
            name_cell.alignment = Alignment(indent=1)
            
            # Debe y Haber (columnas 3 y 4)
            d = ws_resultados.cell(row=row_num, column=3, value=float(debe))
            h = ws_resultados.cell(row=row_num, column=4, value=float(haber))
            
            if cuenta.grupo.naturaleza == 'Ingreso':
                h.fill = haber_fill
                d.fill = debe_fill
            else:  # Gasto
                d.fill = debe_fill
                h.fill = haber_fill
            
            # Total = Haber - Debe (columna 5)
            ws_resultados.cell(row=row_num, column=5, value=f"=D{row_num}-C{row_num}")
            row_num += 1
    
    # Ajustar ancho de columnas
    ws_resultados.column_dimensions['A'].width = 15  # Código
    ws_resultados.column_dimensions['B'].width = 30  # Cuenta
    ws_resultados.column_dimensions['C'].width = 15  # Debe
    ws_resultados.column_dimensions['D'].width = 15  # Haber
    ws_resultados.column_dimensions['E'].width = 15  # Total
    
    # Guardar archivo
    wb.save(output_path)
    print(f"Excel de prueba generado: {output_path}")
    return output_path


def validar_excel_prueba(catalogo, excel_path, empresa, anio=2024):
    """
    Valida el Excel de prueba usando la misma lógica que la vista.
    """
    from openpyxl import load_workbook
    from decimal import Decimal
    
    print(f"\nValidando Excel: {excel_path}")
    print(f"Empresa: {empresa.razon_social}")
    print(f"Año: {anio}")
    
    # Crear o obtener período
    periodo, _ = Periodo.objects.get_or_create(
        empresa=empresa,
        anio=anio,
        mes=None
    )
    
    # Leer Excel
    wb = load_workbook(excel_path, data_only=True)
    filas = []
    errores = []
    creados = 0
    balances_por_tipo = {}
    
    # Leer hoja BalanceGeneral
    if 'BalanceGeneral' in wb.sheetnames:
        ws_balance = wb['BalanceGeneral']
        for row in ws_balance.iter_rows(min_row=3, values_only=True):
            # Saltar filas vacías o que sean títulos de bloque
            if not row[0] or not row[1]:
                continue
            # Saltar si es un subtotal (no tiene código numérico)
            try:
                codigo = str(row[0]).strip()
                # Si el código no parece un código de cuenta, saltar
                if not codigo or codigo.startswith('Total') or not codigo[0].isdigit():
                    continue
            except:
                continue
            
            filas.append({
                'codigo': codigo,
                'nombre': str(row[1]).strip() if row[1] else '',
                'grupo': '',
                'naturaleza': '',
                'tipo_estado': 'BAL',
                'debe': str(row[2]) if row[2] is not None else '0',
                'haber': str(row[3]) if row[3] is not None else '0'
            })
    
    # Leer hoja EstadoResultados
    if 'EstadoResultados' in wb.sheetnames:
        ws_resultados = wb['EstadoResultados']
        for row in ws_resultados.iter_rows(min_row=3, values_only=True):
            # Saltar filas vacías o que sean subtotales
            if not row[0] or not row[1]:
                continue
            # Saltar si es un subtotal (no tiene código numérico)
            try:
                codigo = str(row[0]).strip()
                # Si el código no parece un código de cuenta, saltar
                if not codigo or not codigo[0].isdigit():
                    continue
            except:
                continue
            
            # Ahora la estructura es: Código, Cuenta, Debe, Haber, Total
            filas.append({
                'codigo': codigo,
                'nombre': str(row[1]).strip() if row[1] else '',
                'grupo': '',  # Se obtendrá del catálogo
                'naturaleza': '',  # Se obtendrá del catálogo
                'tipo_estado': 'RES',
                'debe': str(row[2]) if row[2] is not None else '0',  # Columna 3 (índice 2)
                'haber': str(row[3]) if row[3] is not None else '0'  # Columna 4 (índice 3)
            })
    
    # Procesar filas
    for i, row in enumerate(filas, start=1):
        try:
            codigo = row.get('codigo', '').strip()
            nombre = row.get('nombre', '').strip()
            tipo_estado = row.get('tipo_estado', 'BAL').strip().upper()
            
            # Convertir debe y haber de forma segura
            try:
                debe_str = str(row.get('debe', '0') or '0').strip()
                debe = Decimal(debe_str) if debe_str else Decimal('0')
            except (ValueError, TypeError):
                debe = Decimal('0')
            
            try:
                haber_str = str(row.get('haber', '0') or '0').strip()
                haber = Decimal(haber_str) if haber_str else Decimal('0')
            except (ValueError, TypeError):
                haber = Decimal('0')
            
            if not codigo:
                errores.append(f"Fila {i}: Falta código de cuenta")
                continue
            
            # Buscar cuenta en el catálogo
            cuenta = Cuenta.objects.filter(
                grupo__catalogo=catalogo,
                codigo=codigo
            ).select_related('grupo').first()
            
            if not cuenta:
                errores.append(f"Fila {i}: Cuenta {codigo} no encontrada en el catálogo")
                continue
            
            # Crear o obtener balance según tipo
            if tipo_estado not in balances_por_tipo:
                balance, _ = Balance.objects.get_or_create(
                    empresa=empresa,
                    periodo=periodo,
                    tipo_balance=tipo_estado
                )
                balances_por_tipo[tipo_estado] = balance
            else:
                balance = balances_por_tipo[tipo_estado]
            
            # Crear o actualizar detalle de balance
            detalle, created = BalanceDetalle.objects.get_or_create(
                balance=balance,
                cuenta=cuenta,
                defaults={'debe': debe, 'haber': haber}
            )
            if not created:
                detalle.debe = debe
                detalle.haber = haber
                detalle.save()
            
            creados += 1
            
        except Exception as e:
            import traceback
            error_msg = f"Fila {i}: {str(e)}"
            if hasattr(e, '__class__'):
                error_msg += f" [{e.__class__}]"
            errores.append(error_msg)
            # Solo mostrar el traceback completo para el primer error
            if len(errores) == 1:
                print(f"\nError detallado en fila {i}:")
                print(traceback.format_exc())
    
    # Mostrar resultados
    print(f"\n=== RESULTADOS DE VALIDACIÓN ===")
    print(f"Registros procesados: {creados}")
    print(f"Errores encontrados: {len(errores)}")
    
    if errores:
        print("\nErrores:")
        for error in errores[:10]:  # Mostrar solo los primeros 10
            print(f"  - {error}")
        if len(errores) > 10:
            print(f"  ... y {len(errores) - 10} errores más")
    else:
        print("✓ Validación exitosa!")
    
    # Mostrar balances creados
    print(f"\nBalances creados: {len(balances_por_tipo)}")
    for tipo, balance in balances_por_tipo.items():
        detalles = BalanceDetalle.objects.filter(balance=balance).count()
        print(f"  - {tipo}: {detalles} detalles")
    
    return creados, errores


def crear_catalogo_por_defecto(empresa):
    """
    Crea un catálogo por defecto usando CUENTAS_BASE si la empresa no tiene catálogo.
    """
    from stela.services.plantillas import CUENTAS_BASE
    
    print(f"Creando catálogo por defecto para {empresa.razon_social}...")
    
    # Crear o obtener catálogo
    catalogo, created = Catalogo.objects.get_or_create(empresa=empresa)
    
    if not created:
        # Si ya existe, verificar si tiene cuentas
        total_cuentas = Cuenta.objects.filter(grupo__catalogo=catalogo).count()
        if total_cuentas > 0:
            print(f"El catálogo ya existe y tiene {total_cuentas} cuentas")
            return catalogo
    
    print("Creando grupos y cuentas desde CUENTAS_BASE...")
    creados = 0
    
    # Agrupar cuentas por grupo para crear grupos primero
    grupos_dict = {}
    for cuenta_data in CUENTAS_BASE:
        grupo_nombre = cuenta_data['grupo']
        if grupo_nombre not in grupos_dict:
            grupos_dict[grupo_nombre] = []
        grupos_dict[grupo_nombre].append(cuenta_data)
    
    # Crear grupos
    for grupo_nombre, cuentas_data in grupos_dict.items():
        # Obtener naturaleza de la primera cuenta del grupo
        naturaleza = cuentas_data[0]['naturaleza']
        
        grupo, _ = GrupoCuenta.objects.get_or_create(
            catalogo=catalogo,
            nombre=grupo_nombre,
            defaults={'naturaleza': naturaleza}
        )
        
        # Crear cuentas del grupo
        for cuenta_data in cuentas_data:
            cuenta, _ = Cuenta.objects.get_or_create(
                grupo=grupo,
                codigo=cuenta_data['codigo'],
                defaults={
                    'nombre': cuenta_data['nombre'],
                    'aparece_en_balance': True,
                    'bg_bloque': cuenta_data.get('bg_bloque', '') or None,
                    'er_bloque': cuenta_data.get('er_bloque', '') or None,
                    'ratio_tag': cuenta_data.get('ratio_tag', '') or None
                }
            )
            creados += 1
    
    print(f"✓ Catálogo creado con {creados} cuentas")
    return catalogo


def main():
    """Función principal"""
    print("=== GENERADOR Y VALIDADOR DE EXCEL DE PRUEBA ===\n")
    
    # Obtener primera empresa
    empresa = Empresa.objects.first()
    if not empresa:
        print("ERROR: No hay empresas en la base de datos")
        return
    
    print(f"Empresa: {empresa.razon_social}")
    
    # Obtener o crear catálogo
    catalogo = Catalogo.objects.filter(empresa=empresa).first()
    if not catalogo:
        print(f"La empresa {empresa.razon_social} no tiene catálogo")
        catalogo = crear_catalogo_por_defecto(empresa)
    else:
        # Verificar si tiene cuentas
        total_cuentas = Cuenta.objects.filter(grupo__catalogo=catalogo).count()
        if total_cuentas == 0:
            print("El catálogo existe pero no tiene cuentas, creando cuentas por defecto...")
            catalogo = crear_catalogo_por_defecto(empresa)
        else:
            print(f"Catálogo existente con {total_cuentas} cuentas")
    
    print(f"Catálogo ID: {catalogo.id_catalogo}")
    
    # Contar cuentas
    total_cuentas = Cuenta.objects.filter(grupo__catalogo=catalogo).count()
    print(f"Total de cuentas en catálogo: {total_cuentas}\n")
    
    if total_cuentas == 0:
        print("ERROR: No se pudieron crear cuentas en el catálogo")
        return
    
    # Generar Excel de prueba
    excel_path = crear_excel_prueba(catalogo, "test_estados_financieros.xlsx")
    
    # Validar Excel
    creados, errores = validar_excel_prueba(catalogo, excel_path, empresa, anio=2024)
    
    print(f"\n=== RESUMEN ===")
    print(f"Archivo generado: {excel_path}")
    print(f"Registros procesados: {creados}")
    print(f"Errores: {len(errores)}")
    
    if len(errores) == 0:
        print("\n✓ TEST EXITOSO: El Excel se validó correctamente!")
    else:
        print(f"\n✗ TEST FALLIDO: Se encontraron {len(errores)} errores")


if __name__ == '__main__':
    main()


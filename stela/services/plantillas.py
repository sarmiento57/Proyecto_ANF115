import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Protection, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from stela.models.catalogo import Catalogo, GrupoCuenta, Cuenta
from stela.models.empresa import Empresa


# Cuentas base obligatorias para la plantilla de catálogo
# Incluye ratio_tags para cálculo automático de ratios
# Todas las cuentas necesarias para calcular los ratios financieros
CUENTAS_BASE = [
    # ===== ACTIVOS CORRIENTES =====
    
    # Efectivo (ratio_tag: EFECTIVO)
    {'codigo': '1101', 'nombre': 'Caja', 'grupo': 'Activo Corriente', 'naturaleza': 'Activo', 'bg_bloque': 'ACTIVO_CORRIENTE', 'er_bloque': '', 'ratio_tag': 'EFECTIVO'},
    {'codigo': '1102', 'nombre': 'Bancos', 'grupo': 'Activo Corriente', 'naturaleza': 'Activo', 'bg_bloque': 'ACTIVO_CORRIENTE', 'er_bloque': '', 'ratio_tag': 'EFECTIVO'},
    {'codigo': '1103', 'nombre': 'Equivalentes de Efectivo / Inversiones a Corto Plazo', 'grupo': 'Activo Corriente', 'naturaleza': 'Activo', 'bg_bloque': 'ACTIVO_CORRIENTE', 'er_bloque': '', 'ratio_tag': 'EFECTIVO'},
    
    # Cuentas por Cobrar (ratio_tag: CUENTAS_POR_COBRAR)
    {'codigo': '1201', 'nombre': 'Cuentas por Cobrar Comerciales', 'grupo': 'Activo Corriente', 'naturaleza': 'Activo', 'bg_bloque': 'ACTIVO_CORRIENTE', 'er_bloque': '', 'ratio_tag': 'CUENTAS_POR_COBRAR'},
    {'codigo': '1202', 'nombre': 'Estimación para Cuentas Incobrables', 'grupo': 'Activo Corriente', 'naturaleza': 'Activo', 'bg_bloque': 'ACTIVO_CORRIENTE', 'er_bloque': '', 'ratio_tag': ''},
    
    # Inventarios (ratio_tag: INVENTARIOS)
    {'codigo': '1301', 'nombre': 'Inventarios', 'grupo': 'Activo Corriente', 'naturaleza': 'Activo', 'bg_bloque': 'ACTIVO_CORRIENTE', 'er_bloque': '', 'ratio_tag': 'INVENTARIOS'},
    
    # Otros Activos Corrientes
    {'codigo': '1401', 'nombre': 'Otros Activos Corrientes', 'grupo': 'Activo Corriente', 'naturaleza': 'Activo', 'bg_bloque': 'ACTIVO_CORRIENTE', 'er_bloque': '', 'ratio_tag': ''},
    
    # ===== ACTIVOS NO CORRIENTES =====
    
    # Propiedad, Planta y Equipo
    {'codigo': '1501', 'nombre': 'Propiedad, Planta y Equipo - Costo', 'grupo': 'Activo No Corriente', 'naturaleza': 'Activo', 'bg_bloque': 'ACTIVO_NO_CORRIENTE', 'er_bloque': '', 'ratio_tag': ''},
    {'codigo': '1502', 'nombre': 'Depreciación Acumulada de PPE', 'grupo': 'Activo No Corriente', 'naturaleza': 'Activo', 'bg_bloque': 'ACTIVO_NO_CORRIENTE', 'er_bloque': '', 'ratio_tag': 'DEPRECIACION'},
    
    # Activos Intangibles (ratio_tag: AMORTIZACION)
    {'codigo': '1601', 'nombre': 'Activos Intangibles', 'grupo': 'Activo No Corriente', 'naturaleza': 'Activo', 'bg_bloque': 'ACTIVO_NO_CORRIENTE', 'er_bloque': '', 'ratio_tag': ''},
    {'codigo': '1602', 'nombre': 'Amortización Acumulada', 'grupo': 'Activo No Corriente', 'naturaleza': 'Activo', 'bg_bloque': 'ACTIVO_NO_CORRIENTE', 'er_bloque': '', 'ratio_tag': 'AMORTIZACION'},
    
    # Otros Activos No Corrientes
    {'codigo': '1701', 'nombre': 'Otros Activos No Corrientes', 'grupo': 'Activo No Corriente', 'naturaleza': 'Activo', 'bg_bloque': 'ACTIVO_NO_CORRIENTE', 'er_bloque': '', 'ratio_tag': ''},
    
    # ===== PASIVOS CORRIENTES =====
    
    {'codigo': '2101', 'nombre': 'Cuentas por Pagar a Proveedores', 'grupo': 'Pasivo Corriente', 'naturaleza': 'Pasivo', 'bg_bloque': 'PASIVO_CORRIENTE', 'er_bloque': '', 'ratio_tag': ''},
    {'codigo': '2102', 'nombre': 'Otras Cuentas por Pagar / Pasivos Acumulados CP', 'grupo': 'Pasivo Corriente', 'naturaleza': 'Pasivo', 'bg_bloque': 'PASIVO_CORRIENTE', 'er_bloque': '', 'ratio_tag': ''},
    {'codigo': '2103', 'nombre': 'Porción Corriente de Préstamos', 'grupo': 'Pasivo Corriente', 'naturaleza': 'Pasivo', 'bg_bloque': 'PASIVO_CORRIENTE', 'er_bloque': '', 'ratio_tag': ''},
    {'codigo': '2104', 'nombre': 'Documentos por Pagar', 'grupo': 'Pasivo Corriente', 'naturaleza': 'Pasivo', 'bg_bloque': 'PASIVO_CORRIENTE', 'er_bloque': '', 'ratio_tag': ''},
    
    # ===== PASIVOS NO CORRIENTES =====
    
    {'codigo': '2501', 'nombre': 'Préstamos y Obligaciones a Largo Plazo', 'grupo': 'Pasivo No Corriente', 'naturaleza': 'Pasivo', 'bg_bloque': 'PASIVO_NO_CORRIENTE', 'er_bloque': '', 'ratio_tag': ''},
    
    # ===== PATRIMONIO =====
    
    {'codigo': '3101', 'nombre': 'Capital Social', 'grupo': 'Patrimonio', 'naturaleza': 'Patrimonio', 'bg_bloque': 'PATRIMONIO', 'er_bloque': '', 'ratio_tag': ''},
    {'codigo': '3102', 'nombre': 'Reservas / Otras Cuentas de Patrimonio', 'grupo': 'Patrimonio', 'naturaleza': 'Patrimonio', 'bg_bloque': 'PATRIMONIO', 'er_bloque': '', 'ratio_tag': ''},
    {'codigo': '3103', 'nombre': 'Utilidades Retenidas / Resultados Acumulados', 'grupo': 'Patrimonio', 'naturaleza': 'Patrimonio', 'bg_bloque': 'PATRIMONIO', 'er_bloque': '', 'ratio_tag': ''},
    
    # ===== ESTADO DE RESULTADOS - INGRESOS =====
    
    # Ventas Netas (ratio_tag: VENTAS_NETAS)
    {'codigo': '4101', 'nombre': 'Ventas', 'grupo': 'Ingresos', 'naturaleza': 'Ingreso', 'bg_bloque': '', 'er_bloque': 'VENTAS_NETAS', 'ratio_tag': 'VENTAS_NETAS'},
    {'codigo': '4102', 'nombre': 'Descuentos sobre Ventas', 'grupo': 'Ingresos', 'naturaleza': 'Ingreso', 'bg_bloque': '', 'er_bloque': 'VENTAS_NETAS', 'ratio_tag': ''},
    {'codigo': '4103', 'nombre': 'Devoluciones sobre Ventas', 'grupo': 'Ingresos', 'naturaleza': 'Ingreso', 'bg_bloque': '', 'er_bloque': 'VENTAS_NETAS', 'ratio_tag': ''},
    
    # Otros Ingresos (ratio_tag: OTROS_INGRESOS)
    {'codigo': '4301', 'nombre': 'Otros Ingresos', 'grupo': 'Ingresos', 'naturaleza': 'Ingreso', 'bg_bloque': '', 'er_bloque': 'OTROS_INGRESOS', 'ratio_tag': 'OTROS_INGRESOS'},
    
    # ===== ESTADO DE RESULTADOS - COSTOS Y GASTOS =====
    
    # Costo de Ventas (ratio_tag: COSTO_VENTAS)
    {'codigo': '5101', 'nombre': 'Costo de Ventas', 'grupo': 'Costos', 'naturaleza': 'Gasto', 'bg_bloque': '', 'er_bloque': 'COSTO_NETO_VENTAS', 'ratio_tag': 'COSTO_VENTAS'},
    {'codigo': '5102', 'nombre': 'Devoluciones sobre Compras', 'grupo': 'Costos', 'naturaleza': 'Gasto', 'bg_bloque': '', 'er_bloque': 'COSTO_NETO_VENTAS', 'ratio_tag': ''},
    
    # Compras (ratio_tag: COMPRAS)
    {'codigo': '5103', 'nombre': 'Compras de Mercancías', 'grupo': 'Costos', 'naturaleza': 'Gasto', 'bg_bloque': '', 'er_bloque': '', 'ratio_tag': 'COMPRAS'},
    
    # Gastos Operativos (ratio_tag: GASTOS_OPERATIVOS)
    {'codigo': '5201', 'nombre': 'Gastos Operativos / Administración', 'grupo': 'Gastos', 'naturaleza': 'Gasto', 'bg_bloque': '', 'er_bloque': 'GASTOS_OPERATIVOS', 'ratio_tag': 'GASTOS_OPERATIVOS'},
    {'codigo': '5202', 'nombre': 'Gastos de Venta', 'grupo': 'Gastos', 'naturaleza': 'Gasto', 'bg_bloque': '', 'er_bloque': 'GASTOS_OPERATIVOS', 'ratio_tag': 'GASTOS_OPERATIVOS'},
    {'codigo': '5203', 'nombre': 'Gastos Generales', 'grupo': 'Gastos', 'naturaleza': 'Gasto', 'bg_bloque': '', 'er_bloque': 'GASTOS_OPERATIVOS', 'ratio_tag': 'GASTOS_OPERATIVOS'},
    
    # Otros Gastos (ratio_tag: OTROS_GASTOS)
    {'codigo': '5301', 'nombre': 'Otros Gastos', 'grupo': 'Gastos', 'naturaleza': 'Gasto', 'bg_bloque': '', 'er_bloque': 'OTROS_GASTOS', 'ratio_tag': 'OTROS_GASTOS'},
    
    # Gasto Financiero (ratio_tag: GASTO_FINANCIERO)
    {'codigo': '5401', 'nombre': 'Gasto Financiero / Intereses', 'grupo': 'Gastos', 'naturaleza': 'Gasto', 'bg_bloque': '', 'er_bloque': 'GASTO_FINANCIERO', 'ratio_tag': 'GASTO_FINANCIERO'},
    
    # Impuesto sobre la Renta (ratio_tag: IMPUESTO_RENTA)
    {'codigo': '5501', 'nombre': 'Impuesto sobre la Renta', 'grupo': 'Gastos', 'naturaleza': 'Gasto', 'bg_bloque': '', 'er_bloque': 'IMPUESTO_SOBRE_LA_RENTA', 'ratio_tag': 'IMPUESTO_RENTA'},
    
    # Depreciación y Amortización (ratio_tag: DEPRECIACION, AMORTIZACION)
    {'codigo': '5601', 'nombre': 'Gasto por Depreciación', 'grupo': 'Gastos', 'naturaleza': 'Gasto', 'bg_bloque': '', 'er_bloque': '', 'ratio_tag': 'DEPRECIACION'},
    {'codigo': '5602', 'nombre': 'Gasto por Amortización', 'grupo': 'Gastos', 'naturaleza': 'Gasto', 'bg_bloque': '', 'er_bloque': '', 'ratio_tag': 'AMORTIZACION'},
    
    # Servicio de Deuda (ratio_tag: SERVICIO_DEUDA)
    {'codigo': '5701', 'nombre': 'Pago de Capital de Préstamos / Servicio de Deuda', 'grupo': 'Pasivo Corriente', 'naturaleza': 'Pasivo', 'bg_bloque': 'PASIVO_CORRIENTE', 'er_bloque': '', 'ratio_tag': 'SERVICIO_DEUDA'},
]


def generar_plantilla_catalogo_csv():
    """
    DEPRECATED: Esta función está deprecada. Usa generar_plantilla_catalogo_excel en su lugar.
    Genera CSV con cuentas base para catálogo (solo para compatibilidad hacia atrás).
    """
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Encabezados
    writer.writerow(['codigo', 'nombre', 'grupo', 'naturaleza', 'bg_bloque', 'er_bloque', 'ratio_tag'])
    
    # Cuentas base
    for cuenta in CUENTAS_BASE:
        writer.writerow([
            cuenta['codigo'],
            cuenta['nombre'],
            cuenta['grupo'],
            cuenta['naturaleza'],
            cuenta.get('bg_bloque', ''),
            cuenta.get('er_bloque', ''),
            cuenta.get('ratio_tag', '')
        ])
    
    output.seek(0)
    return output.getvalue()


def generar_plantilla_catalogo_excel():
    """Genera Excel con cuentas base para catálogo, con leyenda y protección."""
    wb = Workbook()
    
    # ===== Hoja de leyenda =====
    ws_info = wb.active
    ws_info.title = "Leyenda"
    ws_info["A1"] = "Plantilla de Catálogo - Leyenda"
    ws_info["A1"].font = Font(bold=True, size=14)
    ws_info.append([])
    ws_info.append(["BG Bloque", "Sección del Balance General donde se mostrará la cuenta (ACTIVO_CORRIENTE, PASIVO_CORRIENTE, PATRIMONIO, etc.)."])
    ws_info.append(["ER Bloque", "Subtítulo del Estado de Resultados al que suma la cuenta (VENTAS_NETAS, COSTO_NETO_VENTAS, GASTOS_OPERATIVOS, etc.)."])
    ws_info.append(["ratio_tag", "Etiqueta interna que el autogenerador de ratios espera. Se puede usar '-TAG' para restar (ej. -VENTAS_NETAS)."])
    ws_info.append([])
    ws_info.append(["Nota", "Las columnas bloqueadas (Código, Nombre, Grupo, Naturaleza) no deben editarse. Añade nuevas cuentas debajo si lo necesitas."])
    ws_info.append(["", "Las columnas editables son: BG Bloque, ER Bloque y Ratio Tag."])
    
    # Estilo para la leyenda
    for row in range(1, 8):
        if row == 1:
            ws_info.cell(row=row, column=1).fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        elif row >= 3 and row <= 5:
            ws_info.cell(row=row, column=1).font = Font(bold=True)
            ws_info.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
    
    for col in range(1, 3):
        ws_info.column_dimensions[get_column_letter(col)].width = 55
    
    # ===== Hoja de catálogo =====
    ws = wb.create_sheet("Catálogo")
    
    headers = ['Código', 'Nombre', 'Grupo', 'Naturaleza', 'BG Bloque', 'ER Bloque', 'Ratio Tag']
    
    # Estilo para encabezados
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.protection = Protection(locked=True)
    
    # Cuentas base
    row_idx = 2
    for cuenta in CUENTAS_BASE:
        # Columnas bloqueadas (no editables)
        ws.cell(row=row_idx, column=1, value=cuenta['codigo']).protection = Protection(locked=True)
        ws.cell(row=row_idx, column=2, value=cuenta['nombre']).protection = Protection(locked=True)
        ws.cell(row=row_idx, column=3, value=cuenta['grupo']).protection = Protection(locked=True)
        ws.cell(row=row_idx, column=4, value=cuenta['naturaleza']).protection = Protection(locked=True)
        
        # Columnas editables
        ws.cell(row=row_idx, column=5, value=cuenta.get('bg_bloque', '')).protection = Protection(locked=False)
        ws.cell(row=row_idx, column=6, value=cuenta.get('er_bloque', '')).protection = Protection(locked=False)
        ws.cell(row=row_idx, column=7, value=cuenta.get('ratio_tag', '')).protection = Protection(locked=False)
        
        row_idx += 1
    
    # Ajustar ancho de columnas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 25
    
    # Proteger hoja (permitir insertar filas pero no columnas)
    ws.protection.sheet = True
    ws.protection.insertRows = True  # Permitir insertar filas
    ws.protection.insertColumns = False  # No permitir insertar columnas
    ws.protection.formatCells = False  # No permitir formatear celdas protegidas
    ws.protection.formatRows = False
    ws.protection.formatColumns = False
    ws.protection.deleteRows = True  # Permitir eliminar filas
    ws.protection.deleteColumns = False  # No permitir eliminar columnas
    ws.protection.enable()
    
    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generar_plantilla_estados_csv(catalogo=None):
    """
    DEPRECATED: Esta función está deprecada. Usa generar_plantilla_estados_excel en su lugar.
    Genera CSV para estados financieros basado en el catálogo (solo para compatibilidad hacia atrás).
    Formato esperado: codigo,nombre,grupo,naturaleza,bg_bloque,er_bloque,tipo_estado,debe,haber
    tipo_estado: BAL (Balance General) o RES (Estado de Resultados)
    Si no se proporciona catálogo, usa CUENTAS_BASE como ejemplo.
    """
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Encabezados (formato completo que espera la vista)
    writer.writerow(['codigo', 'nombre', 'grupo', 'naturaleza', 'bg_bloque', 'er_bloque', 'tipo_estado', 'debe', 'haber'])
    
    if catalogo:
        # Obtener cuentas del catálogo
        grupos = GrupoCuenta.objects.filter(catalogo=catalogo).select_related()
        
        # Balance General - agrupar por bg_bloque
        # Solo incluir cuentas que tengan bg_bloque asignado
        for grupo in grupos:
            if grupo.naturaleza in ('Activo', 'Pasivo', 'Patrimonio'):
                cuentas = Cuenta.objects.filter(grupo=grupo, bg_bloque__isnull=False).exclude(bg_bloque='')
                for cuenta in cuentas:
                    writer.writerow([
                        cuenta.codigo,
                        cuenta.nombre,
                        grupo.nombre,
                        grupo.naturaleza,
                        cuenta.bg_bloque,
                        '',  # er_bloque vacío para Balance General
                        'BAL',  # Balance General
                        '0',    # Debe
                        '0'     # Haber
                    ])
        
        # Estado de Resultados - agrupar por er_bloque
        # Solo incluir cuentas que tengan er_bloque asignado
        # Si no hay cuentas para un bloque, simplemente no se incluye (el subtotal será 0)
        for grupo in grupos:
            if grupo.naturaleza in ('Ingreso', 'Gasto'):
                cuentas = Cuenta.objects.filter(grupo=grupo, er_bloque__isnull=False).exclude(er_bloque='')
                for cuenta in cuentas:
                    writer.writerow([
                        cuenta.codigo,
                        cuenta.nombre,
                        grupo.nombre,
                        grupo.naturaleza,
                        '',  # bg_bloque vacío para Estado de Resultados
                        cuenta.er_bloque,
                        'RES',  # Estado de Resultados
                        '0',    # Debe
                        '0'     # Haber
                    ])
    else:
        # Usar CUENTAS_BASE como ejemplo
        # Balance General
        cuentas_balance = [c for c in CUENTAS_BASE if c.get('bg_bloque')]
        for cuenta in cuentas_balance:
            writer.writerow([
                cuenta['codigo'],
                cuenta['nombre'],
                cuenta['grupo'],
                cuenta['naturaleza'],
                cuenta.get('bg_bloque', ''),
                '',  # er_bloque vacío
                'BAL',  # Balance General
                '0',    # Debe
                '0'     # Haber
            ])
        
        # Estado de Resultados
        cuentas_resultados = [c for c in CUENTAS_BASE if c.get('er_bloque')]
        for cuenta in cuentas_resultados:
            writer.writerow([
                cuenta['codigo'],
                cuenta['nombre'],
                cuenta['grupo'],
                cuenta['naturaleza'],
                '',  # bg_bloque vacío
                cuenta.get('er_bloque', ''),
                'RES',  # Estado de Resultados
                '0',    # Debe
                '0'     # Haber
            ])
    
    output.seek(0)
    return output.getvalue()


def generar_plantilla_estados_excel(catalogo):
    """
    Versión mejorada con leyenda visual, protección de celdas, indentación y bordes.
    Genera Excel con dos hojas (BalanceGeneral y EstadoResultados)
    basado en las cuentas del catálogo, agrupadas automáticamente por bloques.
    """
    wb = Workbook()
    
    # Eliminar hoja por defecto
    wb.remove(wb.active)
    
    # Obtener cuentas del catálogo agrupadas por bloques
    grupos = GrupoCuenta.objects.filter(catalogo=catalogo).select_related()
    
    # Agrupar cuentas por bg_bloque para Balance General
    cuentas_por_bg_bloque = {
        'ACTIVO_CORRIENTE': [],
        'ACTIVO_NO_CORRIENTE': [],
        'PASIVO_CORRIENTE': [],
        'PASIVO_NO_CORRIENTE': [],
        'PATRIMONIO': [],
    }
    
    for grupo in grupos:
        cuentas = Cuenta.objects.filter(grupo=grupo)
        for cuenta in cuentas:
            # Balance General - solo incluir cuentas con bg_bloque asignado
            if cuenta.bg_bloque and cuenta.bg_bloque in cuentas_por_bg_bloque:
                cuentas_por_bg_bloque[cuenta.bg_bloque].append(cuenta)
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True)
    
    # Colores de guía visual para Debe y Haber
    debe_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")  # Verde muy claro
    haber_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")  # Azul muy claro
    subtotal_border = Border(top=Side(style="thin"))
    
    # ========== HOJA 1: BALANCE GENERAL ==========
    ws_balance = wb.create_sheet("BalanceGeneral")
    
    # Leyenda de colores (fila 1)
    ws_balance.append(['LEYENDA:', 'Verde claro = Debe (principal para Activos)', 'Azul claro = Haber (principal para Pasivos/Patrimonio)', '', '', '', '', ''])
    ws_balance.merge_cells('A1:D1')
    ws_balance.cell(row=1, column=1).font = Font(bold=True, size=10)
    ws_balance.cell(row=1, column=1).fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # Amarillo claro
    ws_balance.cell(row=1, column=2).fill = debe_fill
    ws_balance.cell(row=1, column=3).fill = haber_fill
    
    # Encabezados Balance General
    ws_balance.append(['Activos', '', '', '', 'Pasivos y Patrimonio', '', '', ''])
    ws_balance.append(['Código', 'Nombre', 'Debe', 'Haber', 'Código', 'Nombre', 'Debe', 'Haber'])
    
    # Estilo encabezados (ahora filas 2 y 3)
    for row in [2, 3]:
        for col in range(1, 9):
            cell = ws_balance.cell(row=row, column=col)
            if row == 2:
                cell.fill = header_fill
                cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.protection = Protection(locked=True)
    
    nombres_bg_bloques = {
        'ACTIVO_CORRIENTE': 'Activo Corriente',
        'ACTIVO_NO_CORRIENTE': 'Activo No Corriente',
        'PASIVO_CORRIENTE': 'Pasivo Corriente',
        'PASIVO_NO_CORRIENTE': 'Pasivo No Corriente',
        'PATRIMONIO': 'Patrimonio'
    }
    
    row_num_activos = 4  # Empezar después de encabezados (fila 1: leyenda, fila 2: títulos, fila 3: headers)
    row_num_pasivos = 4
    subtotales_activos = []  # Para almacenar rangos de subtotales
    subtotales_pasivos = []  # Para almacenar rangos de subtotales
    
    # Activos (columna izquierda)
    for bloque_key in ['ACTIVO_CORRIENTE', 'ACTIVO_NO_CORRIENTE']:
        if bloque_key in cuentas_por_bg_bloque and cuentas_por_bg_bloque[bloque_key]:
            # Título del bloque
            nombre_bloque = nombres_bg_bloques.get(bloque_key, bloque_key)
            ws_balance.cell(row=row_num_activos, column=1, value=nombre_bloque)
            ws_balance.cell(row=row_num_activos, column=1).font = bold_font
            inicio_bloque = row_num_activos + 1
            row_num_activos += 1
            
            # Cuentas del bloque
            for cuenta in cuentas_por_bg_bloque[bloque_key]:
                ws_balance.cell(row=row_num_activos, column=1, value=cuenta.codigo).protection = Protection(locked=True)
                ws_balance.cell(row=row_num_activos, column=2, value=cuenta.nombre).protection = Protection(locked=True)
                
                # Debe y Haber editables
                d = ws_balance.cell(row=row_num_activos, column=3, value=0)
                d.fill = debe_fill
                d.protection = Protection(locked=False)
                
                h = ws_balance.cell(row=row_num_activos, column=4, value=0)
                h.fill = haber_fill
                h.protection = Protection(locked=False)
                
                row_num_activos += 1
            
            # Subtotal del bloque (fila en negritas con borde)
            fin_bloque = row_num_activos - 1
            if inicio_bloque <= fin_bloque:
                ws_balance.cell(row=row_num_activos, column=1, value=f"Total {nombre_bloque}").font = bold_font
                ws_balance.cell(row=row_num_activos, column=1).protection = Protection(locked=True)
                # Fórmula: suma de Haber - suma de Debe
                formula = f"=SUM(D{inicio_bloque}:D{fin_bloque})-SUM(C{inicio_bloque}:C{fin_bloque})"
                total_cell = ws_balance.cell(row=row_num_activos, column=4, value=formula)
                total_cell.font = bold_font
                total_cell.border = subtotal_border
                total_cell.protection = Protection(locked=True)
                subtotales_activos.append(row_num_activos)
                row_num_activos += 1
    
    # Pasivos y Patrimonio (columna derecha)
    for bloque_key in ['PASIVO_CORRIENTE', 'PASIVO_NO_CORRIENTE', 'PATRIMONIO']:
        if bloque_key in cuentas_por_bg_bloque and cuentas_por_bg_bloque[bloque_key]:
            # Título del bloque
            nombre_bloque = nombres_bg_bloques.get(bloque_key, bloque_key)
            ws_balance.cell(row=row_num_pasivos, column=5, value=nombre_bloque)
            ws_balance.cell(row=row_num_pasivos, column=5).font = bold_font
            inicio_bloque = row_num_pasivos + 1
            row_num_pasivos += 1
            
            # Cuentas del bloque
            for cuenta in cuentas_por_bg_bloque[bloque_key]:
                ws_balance.cell(row=row_num_pasivos, column=5, value=cuenta.codigo).protection = Protection(locked=True)
                ws_balance.cell(row=row_num_pasivos, column=6, value=cuenta.nombre).protection = Protection(locked=True)
                
                d = ws_balance.cell(row=row_num_pasivos, column=7, value=0)
                d.fill = debe_fill
                d.protection = Protection(locked=False)
                
                h = ws_balance.cell(row=row_num_pasivos, column=8, value=0)
                h.fill = haber_fill
                h.protection = Protection(locked=False)
                
                row_num_pasivos += 1
            
            # Subtotal del bloque (fila en negritas con borde)
            fin_bloque = row_num_pasivos - 1
            if inicio_bloque <= fin_bloque:
                ws_balance.cell(row=row_num_pasivos, column=5, value=f"Total {nombre_bloque}").font = bold_font
                ws_balance.cell(row=row_num_pasivos, column=5).protection = Protection(locked=True)
                # Fórmula: suma de Haber - suma de Debe
                formula = f"=SUM(H{inicio_bloque}:H{fin_bloque})-SUM(G{inicio_bloque}:G{fin_bloque})"
                total_cell = ws_balance.cell(row=row_num_pasivos, column=8, value=formula)
                total_cell.font = bold_font
                total_cell.border = subtotal_border
                total_cell.protection = Protection(locked=True)
                subtotales_pasivos.append(row_num_pasivos)
                row_num_pasivos += 1
    
    # Total Activos y Total Pasivos + Patrimonio (si hay subtotales)
    if subtotales_activos:
        row_num_activos += 1
        ws_balance.cell(row=row_num_activos, column=1, value="Total Activos").font = bold_font
        ws_balance.cell(row=row_num_activos, column=1).protection = Protection(locked=True)
        # Fórmula: suma de todos los subtotales de activos
        formula = "=" + "+".join([f"D{r}" for r in subtotales_activos])
        c = ws_balance.cell(row=row_num_activos, column=4, value=formula)
        c.font = bold_font
        c.protection = Protection(locked=True)
    
    if subtotales_pasivos:
        row_num_pasivos += 1
        ws_balance.cell(row=row_num_pasivos, column=5, value="Total Pasivos y Patrimonio").font = bold_font
        ws_balance.cell(row=row_num_pasivos, column=5).protection = Protection(locked=True)
        # Fórmula: suma de todos los subtotales de pasivos
        formula = "=" + "+".join([f"H{r}" for r in subtotales_pasivos])
        c = ws_balance.cell(row=row_num_pasivos, column=8, value=formula)
        c.font = bold_font
        c.protection = Protection(locked=True)
    
    # Ajustar ancho de columnas
    for col in range(1, 9):
        ws_balance.column_dimensions[get_column_letter(col)].width = 20
    
    # Proteger hoja
    ws_balance.protection.sheet = True
    ws_balance.protection.enable()
    
    # ========== HOJA 2: ESTADO DE RESULTADOS ==========
    ws_resultados = wb.create_sheet("EstadoResultados")
    
    # Leyenda de colores (fila 1)
    ws_resultados.append(['LEYENDA:', 'Verde claro = Debe (principal para Gastos)', 'Azul claro = Haber (principal para Ingresos)', ''])
    ws_resultados.merge_cells('A1:D1')
    ws_resultados.cell(row=1, column=1).font = Font(bold=True, size=10)
    ws_resultados.cell(row=1, column=1).fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # Amarillo claro
    ws_resultados.cell(row=1, column=2).fill = debe_fill
    ws_resultados.cell(row=1, column=3).fill = haber_fill
    
    # Encabezados Estado de Resultados
    ws_resultados.append(['Código', 'Cuenta', 'Debe', 'Haber', 'Total'])
    
    # Estilo encabezados (ahora fila 2)
    for col in range(1, 6):
        cell = ws_resultados.cell(row=2, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.protection = Protection(locked=True)
    
    # Obtener cuentas agrupadas por bloque ER
    cuentas_por_bloque_er = {
        'VENTAS_NETAS': [],
        'COSTO_NETO_VENTAS': [],
        'GASTOS_OPERATIVOS': [],
        'OTROS_INGRESOS': [],
        'OTROS_GASTOS': [],
        'GASTO_FINANCIERO': [],
        'IMPUESTO_SOBRE_LA_RENTA': [],
    }
    cuentas_sin_bloque_er = []
    
    for grupo in grupos:
        if grupo.naturaleza in ('Ingreso', 'Gasto'):
            cuentas = Cuenta.objects.filter(grupo=grupo)
            for cuenta in cuentas:
                if cuenta.er_bloque and cuenta.er_bloque in cuentas_por_bloque_er:
                    cuentas_por_bloque_er[cuenta.er_bloque].append(cuenta)
                else:
                    cuentas_sin_bloque_er.append(cuenta)
    
    # Orden de bloques ER en el Estado de Resultados
    orden_bloques_er = [
        'VENTAS_NETAS',
        'COSTO_NETO_VENTAS',
        'GASTOS_OPERATIVOS',
        'OTROS_INGRESOS',
        'OTROS_GASTOS',
        'GASTO_FINANCIERO',
        'IMPUESTO_SOBRE_LA_RENTA'
    ]
    
    # Nombres de bloques ER para mostrar
    nombres_bloques_er = {
        'VENTAS_NETAS': 'Ventas Netas',
        'COSTO_NETO_VENTAS': 'Costo Neto de Ventas',
        'GASTOS_OPERATIVOS': 'Gastos Operativos',
        'OTROS_INGRESOS': 'Otros Ingresos',
        'OTROS_GASTOS': 'Otros Gastos',
        'GASTO_FINANCIERO': 'Gasto Financiero',
        'IMPUESTO_SOBRE_LA_RENTA': 'Impuesto sobre la Renta'
    }
    
    row_num = 3  # Empezar después de encabezados (fila 1: leyenda, fila 2: headers)
    
    # Agregar cuentas por bloque ER con fórmulas, indentación y protección
    subtotal_rows = {}
    for bloque_key in orden_bloques_er:
        cuentas = cuentas_por_bloque_er[bloque_key]
        if not cuentas:
            continue
        
        inicio_bloque = row_num
        
        # Agregar cuentas del bloque con fórmula Total = Haber - Debe
        for cuenta in cuentas:
            # Código (columna 1)
            codigo_cell = ws_resultados.cell(row=row_num, column=1, value=cuenta.codigo)
            codigo_cell.protection = Protection(locked=True)
            
            # Nombre indentado (columna 2)
            name_cell = ws_resultados.cell(row=row_num, column=2, value=cuenta.nombre)
            name_cell.alignment = Alignment(indent=1)
            name_cell.protection = Protection(locked=True)
            
            # Debe y Haber editables (columnas 3 y 4)
            d = ws_resultados.cell(row=row_num, column=3, value=0)
            h = ws_resultados.cell(row=row_num, column=4, value=0)
            
            d.protection = Protection(locked=False)
            h.protection = Protection(locked=False)
            
            # Color según naturaleza
            if cuenta.grupo.naturaleza == 'Ingreso':
                h.fill = haber_fill
                d.fill = debe_fill
            else:  # Gasto
                d.fill = debe_fill
                h.fill = haber_fill
            
            # Total = Haber - Debe (bloqueado, columna 5)
            ws_resultados.cell(row=row_num, column=5, value=f"=D{row_num}-C{row_num}").protection = Protection(locked=True)
            
            row_num += 1
        
        fin_bloque = row_num - 1
        
        # Subtotal del bloque (en negritas con borde)
        nombre_bloque = nombres_bloques_er.get(bloque_key, bloque_key)
        ws_resultados.cell(row=row_num, column=1, value="").protection = Protection(locked=True)  # Código vacío
        ws_resultados.cell(row=row_num, column=2, value=nombre_bloque).font = bold_font
        ws_resultados.cell(row=row_num, column=2).protection = Protection(locked=True)
        ws_resultados.cell(row=row_num, column=3, value="").protection = Protection(locked=True)  # Debe vacío
        ws_resultados.cell(row=row_num, column=4, value="").protection = Protection(locked=True)  # Haber vacío
        
        # Fórmula: suma de Haber - suma de Debe (columna 5)
        sub = ws_resultados.cell(row=row_num, column=5, value=f"=SUM(D{inicio_bloque}:D{fin_bloque})-SUM(C{inicio_bloque}:C{fin_bloque})")
        sub.font = bold_font
        sub.border = subtotal_border
        sub.protection = Protection(locked=True)
        subtotal_rows[bloque_key] = row_num
        row_num += 1
    
    # Agregar cuentas sin bloque ER al final (con fórmula Total)
    if cuentas_sin_bloque_er:
        for cuenta in cuentas_sin_bloque_er:
            # Código (columna 1)
            codigo_cell = ws_resultados.cell(row=row_num, column=1, value=cuenta.codigo)
            codigo_cell.protection = Protection(locked=True)
            
            # Nombre indentado (columna 2)
            name_cell = ws_resultados.cell(row=row_num, column=2, value=cuenta.nombre)
            name_cell.alignment = Alignment(indent=1)
            name_cell.protection = Protection(locked=True)
            
            # Debe y Haber editables (columnas 3 y 4)
            d = ws_resultados.cell(row=row_num, column=3, value=0)
            h = ws_resultados.cell(row=row_num, column=4, value=0)
            
            d.protection = Protection(locked=False)
            h.protection = Protection(locked=False)
            
            # Color según naturaleza
            if cuenta.grupo.naturaleza == 'Ingreso':
                h.fill = haber_fill
                d.fill = debe_fill
            else:  # Gasto
                d.fill = debe_fill
                h.fill = haber_fill
            
            # Total = Haber - Debe (bloqueado, columna 5)
            ws_resultados.cell(row=row_num, column=5, value=f"=D{row_num}-C{row_num}").protection = Protection(locked=True)
            row_num += 1
    
    # Agregar Utilidad Bruta (si hay Ventas Netas y Costo Neto de Ventas)
    if 'VENTAS_NETAS' in cuentas_por_bloque_er and 'COSTO_NETO_VENTAS' in cuentas_por_bloque_er:
        # Buscar filas de subtotales
        ventas_row = None
        costo_row = None
        temp_row = 3  # Empezar después de encabezados
        for bloque_key in orden_bloques_er:
            if bloque_key == 'VENTAS_NETAS' and bloque_key in cuentas_por_bloque_er:
                if cuentas_por_bloque_er[bloque_key]:
                    temp_row += len(cuentas_por_bloque_er[bloque_key])
                    ventas_row = temp_row
                    temp_row += 1
            elif bloque_key == 'COSTO_NETO_VENTAS' and bloque_key in cuentas_por_bloque_er:
                if cuentas_por_bloque_er[bloque_key]:
                    temp_row += len(cuentas_por_bloque_er[bloque_key])
                    costo_row = temp_row
                    temp_row += 1
            elif bloque_key in cuentas_por_bloque_er:
                if cuentas_por_bloque_er[bloque_key]:
                    temp_row += len(cuentas_por_bloque_er[bloque_key]) + 1
        
        if ventas_row and costo_row:
            ws_resultados.cell(row=row_num, column=1, value="Utilidad Bruta")
            ws_resultados.cell(row=row_num, column=1).font = bold_font
            total_col = get_column_letter(4)
            formula_utilidad_bruta = f"={total_col}{ventas_row}-{total_col}{costo_row}"
            ws_resultados.cell(row=row_num, column=4, value=formula_utilidad_bruta)
            ws_resultados.cell(row=row_num, column=4).font = bold_font
            row_num += 1
    
    # Agregar Utilidad Operativa (si hay Utilidad Bruta y Gastos Operativos)
    if 'GASTOS_OPERATIVOS' in cuentas_por_bloque_er:
        gastos_row = None
        temp_row = 3  # Empezar después de encabezados
        for bloque_key in orden_bloques_er:
            if bloque_key == 'GASTOS_OPERATIVOS' and bloque_key in cuentas_por_bloque_er:
                if cuentas_por_bloque_er[bloque_key]:
                    temp_row += len(cuentas_por_bloque_er[bloque_key])
                    gastos_row = temp_row
                    temp_row += 1
            elif bloque_key in cuentas_por_bloque_er:
                if cuentas_por_bloque_er[bloque_key]:
                    temp_row += len(cuentas_por_bloque_er[bloque_key]) + 1
        
        if gastos_row and row_num > 2:  # Si hay Utilidad Bruta calculada
            utilidad_bruta_row = row_num - 1
            ws_resultados.cell(row=row_num, column=1, value="Utilidad Operativa")
            ws_resultados.cell(row=row_num, column=1).font = bold_font
            total_col = get_column_letter(4)
            formula_utilidad_operativa = f"={total_col}{utilidad_bruta_row}-{total_col}{gastos_row}"
            ws_resultados.cell(row=row_num, column=4, value=formula_utilidad_operativa)
            ws_resultados.cell(row=row_num, column=4).font = bold_font
            row_num += 1
    
    # Agregar Utilidad Neta (suma de Utilidad Operativa + Otros Ingresos - Otros Gastos - Gasto Financiero - Impuesto)
    # Esto es más complejo, se puede calcular dinámicamente sumando/restando los subtotales relevantes
    
    # Ajustar ancho de columnas
    ws_resultados.column_dimensions['A'].width = 15  # Código
    ws_resultados.column_dimensions['B'].width = 30  # Cuenta
    ws_resultados.column_dimensions['C'].width = 15  # Debe
    ws_resultados.column_dimensions['D'].width = 15  # Haber
    ws_resultados.column_dimensions['E'].width = 15  # Total
    
    # Proteger hoja
    ws_resultados.protection.sheet = True
    ws_resultados.protection.enable()
    
    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


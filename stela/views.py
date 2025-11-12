from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from accounts.decorators import access_required
import pandas as pd
import numpy as np
from datetime import datetime
from django.shortcuts import redirect
from stela.models.empresa import Empresa
from stela.models.venta import Venta
from django.urls import reverse
from django.contrib import messages

from django.shortcuts import get_object_or_404
from django.contrib import messages
from stela.models.empresa import Empresa
from stela.models.finanzas import Periodo, Balance, BalanceDetalle, ResultadoRatio
from stela.models.catalogo import Catalogo, GrupoCuenta, Cuenta
from stela.services.analisis import analisis_vertical, analisis_horizontal
from stela.services.ratios import calcular_y_guardar_ratios
from stela.services.benchmark import benchmarking_por_ciiu, etiqueta_semaforo
from stela.services.ratios_sector import obtener_comparacion_sector
from stela.services.estados import recalcular_saldos_detalle
from stela.services.plantillas import (
    generar_plantilla_catalogo_csv,
    generar_plantilla_catalogo_excel,
    generar_plantilla_estados_csv,
    generar_plantilla_estados_excel
)
from stela.models.ciiu import Ciiu
from stela.forms import CiiuForm, CatalogoUploadForm, CatalogoManualForm, MapeoCuentaForm
from django.core.paginator import Paginator
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from decimal import Decimal
import csv
import io
from openpyxl import load_workbook

from .forms.EmpresaForm import EmpresaForm,EmpresaEditForm

from .models.empresa import Empresa
from .models.finanzas import RatioDef, Periodo, ResultadoRatio, BalanceDetalle
from .models.catalogo import Catalogo, Cuenta
from django.db.models import F

# Create your views here.
def landing(request):
    return render(request, "stela/landing.html")


@access_required('002')
def dashboard(request):
    """Vista del dashboard con información de empresas y catálogos"""
    user = request.user

    # Mostrar mensaje solo una vez por sesión
    if not request.session.get('bienvenida_mostrada'):
        messages.success(request, f'Bienvenido, {user.first_name} {user.last_name}')
        request.session['bienvenida_mostrada'] = True
        
    # se cambio por el nuevo campo many to many 
    empresas = Empresa.objects.filter(usuario=user).distinct().order_by('razon_social')
    
    # Búsqueda por empresa
    query = request.GET.get('q', '').strip()
    if query:
        empresas = empresas.filter(razon_social__icontains=query)
    
    # Obtener información de catálogos y estados financieros por empresa
    empresas_data = []
    for empresa in empresas:
        catalogo = Catalogo.objects.filter(empresa=empresa).first()
        periodos = Periodo.objects.filter(empresa=empresa).order_by('-anio', '-mes')
        balances = Balance.objects.filter(empresa=empresa).select_related('periodo').order_by('-periodo__anio', '-periodo__mes')
        
        empresas_data.append({
            'empresa': empresa,
            'catalogo': catalogo,
            'tiene_catalogo': catalogo is not None,
            'num_periodos': periodos.count(),
            'num_balances': balances.count(),
            'periodos': periodos[:5],  # Últimos 5 períodos
        })
    
    context = {
        'empresas_data': empresas_data,
        'tiene_empresas': empresas.exists(),
        'query': query,
    }
    
    return render(request, "dashboard/dashboard.html", context)


@access_required('003')
def empresa_detalles(request, empresa_nit):
    """Vista de detalles de una empresa: estados financieros y catálogo"""
    user = request.user
    #se cambio por el nuevo campo many to many
    empresa = get_object_or_404(Empresa.objects.filter(usuario=user), nit=empresa_nit)
    
    # Obtener catálogo de la empresa
    catalogo = Catalogo.objects.filter(empresa=empresa).first()
    cuentas = []
    if catalogo:
        cuentas = Cuenta.objects.filter(grupo__catalogo=catalogo).select_related('grupo').order_by('codigo')
    
    # Obtener períodos que tienen al menos un balance
    periodos = Periodo.objects.filter(
        empresa=empresa,
        balance__isnull=False
    ).distinct().order_by('-anio', '-mes')
    
    # Eliminar períodos vacíos (sin balances)
    periodos_vacios = Periodo.objects.filter(empresa=empresa).exclude(
        id_periodo__in=periodos.values_list('id_periodo', flat=True)
    )
    
    if periodos_vacios.exists():
        # Guardar el conteo antes de eliminar
        num_periodos_vacios = periodos_vacios.count()
        # Eliminar ratios calculados para períodos vacíos primero
        from stela.models.finanzas import ResultadoRatio
        ResultadoRatio.objects.filter(periodo__in=periodos_vacios).delete()
        # Eliminar los períodos vacíos
        periodos_vacios.delete()
        messages.info(request, f'Se eliminaron {num_periodos_vacios} período(s) vacío(s) sin estados financieros.')
    
    # Agrupar balances por período
    balances_por_periodo = {}
    for periodo in periodos:
        balances = Balance.objects.filter(empresa=empresa, periodo=periodo).select_related('periodo')
        balances_por_periodo[periodo] = {
            'balance': balances.filter(tipo_balance='BAL').first(),
            'resultados': balances.filter(tipo_balance='RES').first(),
        }
    
    context = {
        'empresa': empresa,
        'catalogo': catalogo,
        'cuentas': cuentas,
        'tiene_catalogo': catalogo is not None,
        'periodos': periodos,
        'balances_por_periodo': balances_por_periodo,
    }
    return render(request, "dashboard/empresa_detalles.html", context)

@access_required('041', stay_on_page=True)
def crearEmpresa(request):
    # Initialize form variable outside the if block
    form = EmpresaForm()

    if request.method == 'POST':
        form = EmpresaForm(request.POST)

        if form.is_valid():
            try:
                empresa = form.save(commit=False)
                empresa.save()  # Guardar primero para que tenga ID
                empresa.usuario.add(request.user)  # Luego agregar usuario (ManyToMany)
                messages.success(request, f'Empresa {empresa.razon_social} creada exitosamente.')
                return redirect('dashboard')
            except Exception as e:
                messages.error(request, f'Error al crear la empresa: {str(e)}')
                # Log del error para debugging
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f'Error al crear empresa: {e}', exc_info=True)

    context = {
        'form': form
    }
    return render(request, "stela/crear-empresa.html", context)

@access_required('010', stay_on_page=True)
def eliminar_balance(request, balance_id):
    """
    Vista para eliminar un estado financiero (Balance) y todos sus detalles.
    También elimina los ratios calculados para ese período si ya no hay balances.
    """
    from stela.models.finanzas import Balance, BalanceDetalle, ResultadoRatio, Periodo
    
    balance = get_object_or_404(
        Balance.objects.select_related('empresa', 'periodo'),
        pk=balance_id,
        empresa__usuario=request.user
    )
    
    empresa = balance.empresa
    periodo = balance.periodo
    tipo_balance = balance.get_tipo_balance_display()
    
    if request.method == 'POST':
        try:
            # Eliminar el balance (esto eliminará automáticamente todos los BalanceDetalle por CASCADE)
            balance.delete()
            
            # Verificar si el período ya no tiene balances
            balances_restantes = Balance.objects.filter(periodo=periodo).count()
            
            # Si no quedan balances, eliminar también los ratios calculados para ese período
            if balances_restantes == 0:
                ResultadoRatio.objects.filter(empresa=empresa, periodo=periodo).delete()
                messages.info(request, f'Se eliminaron también los ratios calculados para el período {periodo}.')
            
            messages.success(request, f'{tipo_balance} del período {periodo} eliminado correctamente.')
            return redirect('empresa_detalles', empresa_nit=empresa.nit)
        except Exception as e:
            messages.error(request, f'Error al eliminar el estado financiero: {str(e)}')
            return redirect('empresa_detalles', empresa_nit=empresa.nit)
    
    # GET: Mostrar confirmación
    context = {
        'balance': balance,
        'empresa': empresa,
        'periodo': periodo,
        'tipo_balance': tipo_balance,
        'num_detalles': balance.detalles.count()
    }
    return render(request, 'stela/balance/confirmar_eliminar.html', context)

@access_required('042', stay_on_page=True)
def editarEmpresa(request, nit):
    empresa = get_object_or_404(Empresa.objects.filter(usuario=request.user), nit=nit)
    if request.method == 'POST':

        form = EmpresaEditForm(request.POST, instance=empresa)

        if form.is_valid():
            form.save()
            return redirect('dashboard')

    else:
        form = EmpresaEditForm(instance=empresa)

    context = {
        'form': form,
        'empresa': empresa
    }
    return render(request, "stela/editar-empresa.html", context)

@access_required('005')
def tools(request):
    """Vista inicial de herramientas - redirige a tools_finanzas"""
    user = request.user
    # Obtener empresas disponibles para el usuario (igual que en dashboard)
    empresas = Empresa.objects.filter(usuario=user).distinct().order_by('razon_social')
    ctx = {
        'estado': 'RES',
        'empresas': empresas,
        'tiene_empresas': empresas.exists(),
    }
    return render(request, 'tools/tools.html', ctx)

@access_required('040', stay_on_page=True)
def tools_finanzas(request):
    # parámetros: ?nit=...&per_base=ID&per_act=ID
    # Ahora calcula ambos tipos de estado simultáneamente
    user = request.user
    nit = request.GET.get('nit')
    per_base_id = request.GET.get('per_base')
    per_act_id  = request.GET.get('per_act')

    # Obtener empresas disponibles para el usuario (igual que en dashboard)
    empresas = Empresa.objects.filter(usuario=user).distinct().order_by('razon_social')

    ctx = {
        'empresas': empresas,
        'tiene_empresas': empresas.exists(),
    }
    
    if not (nit and per_act_id):
        messages.info(request, "Selecciona empresa y período para consultar.")
        return render(request, "tools/tools.html", ctx)

    # se cambio por el nuevo campo many to many
    empresa = get_object_or_404(Empresa.objects.filter(usuario=request.user), nit=nit)
    p_act   = get_object_or_404(Periodo, pk=per_act_id)

    # CONSULTAR ratios ya calculados desde ResultadoRatio (no calcular)
    # Los ratios se calculan automáticamente al cargar estados financieros
    ratios_guardados = ResultadoRatio.objects.filter(
        empresa=empresa,
        periodo=p_act
    ).select_related('ratio')
    
    # Convertir a formato de diccionario para compatibilidad con código existente
    ratios_dict = {}
    for rr in ratios_guardados:
        clave = rr.ratio.clave
        if clave not in ratios_dict or (rr.valor is not None and ratios_dict[clave]['valor'] is None):
            ratios_dict[clave] = {
                'clave': clave,
                'nombre': rr.ratio.nombre,
                'valor': rr.valor
            }
    
    # Si no hay ratios guardados, intentar calcular en tiempo real como respaldo
    if not ratios_dict:
        try:
            ratios_bal = calcular_y_guardar_ratios(empresa, p_act, tipo_estado='BAL')
            ratios_res = calcular_y_guardar_ratios(empresa, p_act, tipo_estado='RES')
            for r in ratios_bal + ratios_res:
                if r['clave'] not in ratios_dict:
                    ratios_dict[r['clave']] = r
                elif r['valor'] is not None:
                    ratios_dict[r['clave']] = r
            if ratios_dict:
                messages.info(request, "Ratios calculados en tiempo real. Se recomienda cargar estados financieros para calcular ratios automáticamente.")
        except Exception as e:
            messages.warning(request, f"No se encontraron ratios calculados para este período. Error al calcular: {str(e)}")
    
    ratios = list(ratios_dict.values())

    # Análisis Vertical - Estado de Resultados
    vertical_act_res = analisis_vertical(empresa, p_act, 'RES')
    vertical_base_res = []
    horizontal_rows_res = []
    
    # Análisis Vertical - Balance General
    vertical_act_bal = analisis_vertical(empresa, p_act, 'BAL')
    vertical_base_bal = []
    horizontal_rows_bal = []

    # Horizontal si hay base
    if per_base_id:
        p_base = get_object_or_404(Periodo, pk=per_base_id)
        # Estado de Resultados
        horizontal_rows_res = analisis_horizontal(empresa, p_base, p_act, 'RES')
        vertical_base_res = analisis_vertical(empresa, p_base, 'RES')
        # Balance General
        horizontal_rows_bal = analisis_horizontal(empresa, p_base, p_act, 'BAL')
        vertical_base_bal = analisis_vertical(empresa, p_base, 'BAL')

    # Benchmark: Promedio histórico de la misma empresa (todos los períodos)
    ratios_bench = []
    from decimal import Decimal
    from statistics import mean, pstdev
    
    # Obtener todos los períodos de la empresa con ratios calculados
    periodos_empresa = Periodo.objects.filter(
        empresa=empresa,
        resultadoratio__isnull=False
    ).distinct()
    
    for r in ratios:
        # Buscar el ratio definido
        try:
            ratio_def = RatioDef.objects.get(clave=r['clave'])
        except RatioDef.DoesNotExist:
            continue
        
        # Obtener todos los valores históricos de este ratio para la empresa
        valores_historicos = list(
            ResultadoRatio.objects.filter(
                empresa=empresa,
                ratio=ratio_def,
                periodo__in=periodos_empresa,
                valor__isnull=False
            ).values_list('valor', flat=True)
        )
        
        if valores_historicos and len(valores_historicos) > 0:
            # Calcular promedio y desviación estándar
            valores_decimal = [Decimal(str(v)) for v in valores_historicos]
            prom_historico = Decimal(mean(valores_decimal))
            desv_historico = Decimal(pstdev(valores_decimal)) if len(valores_decimal) > 1 else Decimal('0')
            
            # Comparar valor actual con promedio histórico
            valor_actual = r.get('valor')
            if valor_actual is not None:
                # Calcular semáforo comparando valor actual vs promedio histórico
                if desv_historico == 0:
                    sem = 'OK' if valor_actual == prom_historico else ('ALTO' if valor_actual > prom_historico else 'BAJO')
                else:
                    k = Decimal('1')
                    if valor_actual > prom_historico + k * desv_historico:
                        sem = 'ALTO'
                    elif valor_actual < prom_historico - k * desv_historico:
                        sem = 'BAJO'
                    else:
                        sem = 'OK'
            else:
                sem = 'NA'
            
            ratios_bench.append({
                **r,
                'promedio': prom_historico,
                'desv': desv_historico,
                'semaforo': sem
            })
        else:
            # Si no hay valores históricos, mostrar solo el valor actual
            ratios_bench.append({
                **r,
                'promedio': None,
                'desv': None,
                'semaforo': 'NA'
            })

    # Comparación con parámetros de sector (ratios digitados)
    ratios_sector = []
    if empresa.ciiu:
        ratios_sector = obtener_comparacion_sector(empresa, ratios)
    
    # Contar períodos con ratios para validar botón de gráficas
    periodos_con_ratios = Periodo.objects.filter(
        empresa=empresa,
        resultadoratio__isnull=False
    ).distinct().count()
    puede_ver_graficas = periodos_con_ratios >= 3

    ctx.update({
        'empresa': empresa,
        'per_act': p_act,
        'per_base_id': per_base_id,
        # Estado de Resultados
        'vertical_act_res': vertical_act_res,
        'vertical_base_res': vertical_base_res,
        'horizontal_rows_res': horizontal_rows_res,
        # Balance General
        'vertical_act_bal': vertical_act_bal,
        'vertical_base_bal': vertical_base_bal,
        'horizontal_rows_bal': horizontal_rows_bal,
        # Ratios
        'ratios_rows': ratios_bench or ratios,
        'ratios_sector': ratios_sector,
        'puede_ver_graficas': puede_ver_graficas,
        'periodos_con_ratios': periodos_con_ratios,
    })
    return render(request, "tools/tools.html", ctx)

@access_required('006')
def projections(request):
    user = request.user
    #se cambio por el nuevo campo many to many
    empresas = Empresa.objects.filter(usuario=user).distinct()
    empresa_seleccionada = None
    ventas = []
    proyecciones_minimos = []
    proyecciones_porcentuales = []
    proyecciones_incremento_absoluto = []

    if request.method == "POST" and "archivo_excel" in request.FILES:
        empresa_nit = request.POST.get("empresa")
        # se cambio por el nuevo campo many to many
        empresa_seleccionada = get_object_or_404(Empresa.objects.filter(usuario=user), nit=empresa_nit)
        archivo = request.FILES["archivo_excel"]

        # eliminar registros anteriores
        Venta.objects.filter(empresa=empresa_seleccionada).delete()

        # leer el excel
        try:
            df = pd.read_excel(archivo)
        except Exception as e:
            messages.error(request, f"Error al leer el archivo Excel: {e}")
            return redirect(f"{reverse('projections')}?empresa={empresa_nit}")

        # normalizar nombres de columnas
        df.columns = df.columns.str.strip().str.lower()

        # buscar las columnas de mes y valor
        col_mes = next((c for c in df.columns if "mes" in str(c).lower() or "fecha" in str(c).lower()), None)
        col_valor = next((c for c in df.columns if "valor" in str(c).lower() or "monto" in str(c).lower() or "venta" in str(c).lower()), None)

        # valida enr que existan las columnas
        if not col_mes or not col_valor:
            messages.error(
                request,
                "El archivo debe tener columnas llamadas 'Mes' y 'Valor'.",
            )
            return redirect(f"{reverse('projections')}?empresa={empresa_nit}")

        # si todo esta verygod validamos y guardamos
        try:
            for i, row in df.iterrows():
                # mes
                try:
                    fecha = pd.to_datetime(row[col_mes], errors="raise")
                except Exception:
                    messages.error(
                        request,
                        f"Error en la fila {i + 2}: la columna 'Mes' tiene un valor no válido ('{row[col_mes]}'). "
                        "Debe ser una fecha como 2024-01-01.",
                    )
                    return redirect(f"{reverse('projections')}?empresa={empresa_nit}")

                # valor
                try:
                    valor = float(row[col_valor])
                except Exception:
                    messages.error(
                        request,
                        f"Error en la fila {i + 2}: la columna 'Valor' tiene un valor no numérico ('{row[col_valor]}').",
                    )
                    return redirect(f"{reverse('projections')}?empresa={empresa_nit}")

                # guarda en ventas
                Venta.objects.create(
                    empresa=empresa_seleccionada,
                    mes_venta=fecha,
                    saldo_venta=valor,
                    anio=fecha.year,
                    proyeccion=False
                )

            # verygud
            messages.success(
                request,
                f"Ventas guardadas correctamente para la empresa {empresa_seleccionada.razon_social}.",
            )

        except Exception as e:
            messages.error(request, f"Error inesperado al procesar el archivo: {e}")
            return redirect(f"{reverse('projections')}?empresa={empresa_nit}")

            # Redirigir para recargar datos y mostrar mensaje
            return redirect(f"{reverse('projections')}?empresa={empresa_nit}")
        
    # procesar la empresa seleccionada
    empresa_nit = request.GET.get("empresa")
    if empresa_nit:
        # se cambio por el nuevo campo many to many
        empresa_seleccionada = get_object_or_404(Empresa.objects.filter(usuario=user), nit=empresa_nit)

        # ventas reales
        ventas = Venta.objects.filter(
            empresa=empresa_seleccionada, proyeccion=False
        ).order_by("mes_venta")

        # minimos cuadrados
        y = np.array([v.saldo_venta for v in ventas])
        n = len(y)
        if n > 1:
            x = np.arange(1, n + 1)
            sum_x = np.sum(x)
            sum_y = np.sum(y)
            sum_xy = np.sum(x * y)
            sum_x2 = np.sum(x**2)

            b = (n * sum_xy - sum_x * sum_y) / (n * sum_x2 - sum_x**2)
            a = (sum_y - b * sum_x) / n

            x_future = np.arange(n + 1, n + 13)
            y_future = a + b * x_future
            anio_base = ventas.last().anio

            proyecciones_minimos = [
                {
                    "mes_proyectado": datetime(anio_base + 1, i + 1, 1),
                    "valor_proyectado": y_val,
                }
                for i, y_val in enumerate(y_future)
            ]

        # incremento porcentual
        valores = [v.saldo_venta for v in ventas]
        if len(valores) > 1:
            incrementos = [
                (valores[i] - valores[i - 1]) / valores[i - 1]
                for i in range(1, len(valores))
            ]
            prom_incremento = np.mean(incrementos)
            ultimo_valor = valores[-1]
            anio_base = ventas.last().anio

            for mes in range(1, 13):
                nuevo_valor = ultimo_valor * (1 + prom_incremento)
                ultimo_valor = nuevo_valor
                proyecciones_porcentuales.append(
                    {
                        "mes_proyectado": datetime(anio_base + 1, mes, 1),
                        "valor_proyectado": nuevo_valor,
                    }
                )

        # incremento absoluto
        if len(valores) > 1:
            incrementos_abs = [
                (valores[i] - valores[i - 1]) for i in range(1, len(valores))
            ]
            prom_incremento_abs = np.mean(incrementos_abs)
            ultimo_valor = valores[-1]
            anio_base = ventas.last().anio

            for mes in range(1, 13):
                nuevo_valor = ultimo_valor + prom_incremento_abs
                ultimo_valor = nuevo_valor
                proyecciones_incremento_absoluto.append(
                    {
                        "mes_proyectado": datetime(anio_base + 1, mes, 1),
                        "valor_proyectado": nuevo_valor,
                    }
                )
                
    contexto = {
        "empresas": empresas,
        "empresa_seleccionada": empresa_seleccionada,
        "ventas": ventas,
        "proyecciones": proyecciones_minimos,
        "proyecciones_porcentuales": proyecciones_porcentuales,
        "proyecciones_incremento_absoluto": proyecciones_incremento_absoluto,
    }
    return render(request, "projections/projection.html", contexto)

#filtro para las proyecciones
def money(value):
    try:
        return "${:,.2f}".format(float(value))
    except (ValueError, TypeError):
        return "$0.00"


# ========== VISTAS CRUD PARA CIIU (CATÁLOGOS) ==========

@access_required('007')
def ciiu_list(request):
    """
    Lista todos los códigos CIIU con paginación y búsqueda.
    """
    query = request.GET.get('q', '').strip()
    ciiu_list = Ciiu.objects.all().order_by('codigo')
    
    # Búsqueda por código o descripción
    if query:
        ciiu_list = ciiu_list.filter(
            codigo__icontains=query
        ) | ciiu_list.filter(
            descripcion__icontains=query
        )
    
    # Paginación
    paginator = Paginator(ciiu_list, 25)  # 25 por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Contar empresas asociadas para cada CIIU
    for ciiu in page_obj:
        ciiu.empresas_count = Empresa.objects.filter(ciiu=ciiu).count()
    
    context = {
        'page_obj': page_obj,
        'query': query,
    }
    return render(request, 'stela/catalogo/list.html', context)


@access_required('036', stay_on_page=True)
def ciiu_create(request):
    """
    Crea un nuevo código CIIU.
    """
    if request.method == 'POST':
        form = CiiuForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Código CIIU creado correctamente.')
            return redirect('ciiu_list')
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else:
        form = CiiuForm()
    
    context = {
        'form': form,
        'titulo': 'Crear Código CIIU'
    }
    return render(request, 'stela/catalogo/create.html', context)


@access_required('037', stay_on_page=True)
def ciiu_update(request, codigo):
    """
    Edita un código CIIU existente.
    """
    ciiu = get_object_or_404(Ciiu, codigo=codigo)
    
    if request.method == 'POST':
        form = CiiuForm(request.POST, instance=ciiu)
        if form.is_valid():
            form.save()
            messages.success(request, 'Código CIIU actualizado correctamente.')
            return redirect('ciiu_list')
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else:
        form = CiiuForm(instance=ciiu)
    
    # Contar empresas asociadas
    empresas_count = Empresa.objects.filter(ciiu=ciiu).count()
    hijos_count = ciiu.hijos.count()
    
    context = {
        'form': form,
        'ciiu': ciiu,
        'titulo': 'Editar Código CIIU',
        'empresas_count': empresas_count,
        'hijos_count': hijos_count
    }
    return render(request, 'stela/catalogo/update.html', context)


@access_required('038', stay_on_page=True)
def ciiu_delete(request, codigo):
    """
    Elimina un código CIIU (con validación de empresas e hijos).
    """
    ciiu = get_object_or_404(Ciiu, codigo=codigo)
    
    # Validar si tiene empresas asociadas
    empresas_count = Empresa.objects.filter(ciiu=ciiu).count()
    hijos_count = ciiu.hijos.count()
    
    if request.method == 'POST':
        # Verificar nuevamente antes de eliminar
        if Empresa.objects.filter(ciiu=ciiu).exists():
            messages.error(
                request,
                f'No se puede eliminar el código CIIU porque tiene {empresas_count} empresa(s) asociada(s).'
            )
            return redirect('ciiu_list')
        
        if ciiu.hijos.exists():
            messages.error(
                request,
                f'No se puede eliminar el código CIIU porque tiene {hijos_count} código(s) hijo(s).'
            )
            return redirect('ciiu_list')
        
        ciiu.delete()
        messages.success(request, 'Código CIIU eliminado correctamente.')
        return redirect('ciiu_list')
    
    context = {
        'ciiu': ciiu,
        'empresas_count': empresas_count,
        'hijos_count': hijos_count
    }
    return render(request, 'stela/catalogo/delete.html', context)


# ========== VISTAS PARA CATÁLOGO DE CUENTAS ==========

@access_required('008', stay_on_page=True)
def catalogo_upload_csv(request):
    """
    Vista para cargar catálogo desde CSV.
    Maneja los pasos 1, 2 y 3 del asistente:
    - Paso 1: Selección de empresa
    - Paso 2: Carga solo el catálogo (cuentas)
    - Paso 3: Carga estados financieros (debe/haber)
    """
    paso = request.GET.get('paso', '1')  # Por defecto paso 1 (Selecciona Empresa)
    empresa_nit_param = request.GET.get('empresa', '')  # Empresa desde parámetro URL
    catalogo_id_param = request.GET.get('catalogo_id', '')  # Catálogo desde parámetro URL
    
    if request.method == 'POST':
        empresa_nit = request.POST.get('empresa')
        anio = request.POST.get('anio')  # Año para el período de estados financieros
        archivo = request.FILES.get('archivo')
        
        if not empresa_nit or not archivo:
            messages.error(request, 'Faltan datos requeridos')
            return redirect('catalogo_upload')
        
        try:
            # se cambio por el nuevo campo many to many
            empresa = get_object_or_404(Empresa.objects.filter(usuario=request.user), nit=empresa_nit)
            
            # Crear o obtener catálogo (único por empresa, sin año)
            catalogo, created = Catalogo.objects.get_or_create(
                empresa=empresa
            )
            
            # Solo crear período si estamos en paso 3 (estados financieros) y hay año
            periodo = None
            if paso == '3' and anio:
                anio = int(anio)
                periodo, _ = Periodo.objects.get_or_create(
                    empresa=empresa,
                    anio=anio,
                    mes=None
                )
            
            # Detectar tipo de archivo y leer
            nombre_archivo = archivo.name.lower()
            es_excel = nombre_archivo.endswith(('.xlsx', '.xls'))
            
            creados = 0
            errores = []
            
            # Diccionario para agrupar por tipo de estado
            balances_por_tipo = {}
            
            if es_excel:
                # Procesar Excel
                wb = load_workbook(archivo, data_only=True)
                filas = []
                
                if paso == '2':
                    # Paso 2: Catálogo (una sola hoja "Catálogo")
                    hoja_nombre = 'Catálogo' if 'Catálogo' in wb.sheetnames else wb.active.title
                    ws = wb[hoja_nombre]
                    for row in ws.iter_rows(min_row=2, values_only=True):  # Empezar desde fila 2 (después de encabezados)
                        if row[0] and row[1]:  # Si hay código y nombre
                            filas.append({
                                'codigo': str(row[0]).strip() if row[0] else '',
                                'nombre': str(row[1]).strip() if row[1] else '',
                                'grupo': str(row[2]).strip() if len(row) > 2 and row[2] else '',
                                'naturaleza': str(row[3]).strip() if len(row) > 3 and row[3] else '',
                                'bg_bloque': str(row[4]).strip() if len(row) > 4 and row[4] else '',
                                'er_bloque': str(row[5]).strip() if len(row) > 5 and row[5] else '',
                                'ratio_tag': str(row[6]).strip() if len(row) > 6 and row[6] else '',
                                'tipo_estado': '',  # No aplica en paso 2
                                'debe': '0',
                                'haber': '0'
                            })
                else:
                    # Paso 3: Estados financieros (dos hojas)
                    # Leer hoja BalanceGeneral
                    if 'BalanceGeneral' in wb.sheetnames:
                        ws_balance = wb['BalanceGeneral']
                        # Empezar desde fila 4 (fila 1: leyenda, fila 2: títulos, fila 3: encabezados, fila 4+: datos)
                        # Leer hasta el final de la hoja (sin límite máximo de filas)
                        max_row = ws_balance.max_row
                        for row in ws_balance.iter_rows(min_row=4, max_row=max_row, values_only=True):
                            # Leer columna izquierda (Activos): columnas 0-3 (Código, Nombre, Debe, Haber)
                            if len(row) > 1 and row[0] and row[1]:
                                try:
                                    codigo = str(row[0]).strip()
                                    # Saltar si es un título de bloque, subtotal o no tiene código numérico
                                    if codigo and not codigo.startswith('Total') and codigo[0].isdigit():
                                        filas.append({
                                            'codigo': codigo,
                                            'nombre': str(row[1]).strip() if row[1] else '',
                                            'grupo': '',  # Se obtendrá del catálogo
                                            'naturaleza': '',  # Se obtendrá del catálogo
                                            'tipo_estado': 'BAL',
                                            'debe': str(row[2]) if len(row) > 2 and row[2] is not None else '0',
                                            'haber': str(row[3]) if len(row) > 3 and row[3] is not None else '0'
                                        })
                                except Exception as e:
                                    # Log del error para debugging si es necesario
                                    pass
                            
                            # Leer columna derecha (Pasivos y Patrimonio): columnas 4-7 (Código, Nombre, Debe, Haber)
                            # En la plantilla, las columnas son: 5=Código, 6=Nombre, 7=Debe, 8=Haber (índices 4,5,6,7)
                            # Verificar que la fila tenga suficientes columnas
                            if len(row) > 5:
                                try:
                                    # La columna 5 (índice 4) puede tener código o título de bloque
                                    # La columna 6 (índice 5) tiene el nombre
                                    codigo_val = row[4]
                                    nombre_val = row[5]
                                    
                                    # Solo procesar si hay un código válido (no None, no vacío)
                                    if codigo_val is not None:
                                        codigo = str(codigo_val).strip()
                                        nombre = str(nombre_val).strip() if nombre_val else ''
                                        
                                        # Saltar títulos de bloque (como "Patrimonio", "Pasivo Corriente", etc.)
                                        # Estos no empiezan con dígito
                                        # Saltar subtotales que empiezan con "Total"
                                        if codigo and not codigo.startswith('Total') and codigo[0].isdigit() and nombre:
                                            filas.append({
                                                'codigo': codigo,
                                                'nombre': nombre,
                                                'grupo': '',  # Se obtendrá del catálogo
                                                'naturaleza': '',  # Se obtendrá del catálogo
                                                'tipo_estado': 'BAL',
                                                'debe': str(row[6]) if len(row) > 6 and row[6] is not None else '0',
                                                'haber': str(row[7]) if len(row) > 7 and row[7] is not None else '0'
                                            })
                                except Exception as e:
                                    # Log del error para debugging si es necesario
                                    pass
                    
                    # Leer hoja EstadoResultados
                    if 'EstadoResultados' in wb.sheetnames:
                        ws_resultados = wb['EstadoResultados']
                        # Empezar desde fila 3 (fila 1: leyenda, fila 2: encabezados, fila 3+: datos)
                        for row in ws_resultados.iter_rows(min_row=3, values_only=True):
                            # Saltar filas vacías o que sean subtotales
                            if not row[0] or not row[1]:
                                continue
                            # Saltar si es un subtotal o título de bloque (no tiene código numérico)
                            try:
                                codigo = str(row[0]).strip()
                                # Si el código no parece un código de cuenta, saltar
                                if not codigo or codigo.startswith('Total') or not codigo[0].isdigit():
                                    continue
                            except:
                                continue
                            
                            # Estructura: Código, Cuenta, Debe, Haber, Total
                            filas.append({
                                'codigo': codigo,
                                'nombre': str(row[1]).strip() if row[1] else '',
                                'grupo': '',  # Se obtendrá del catálogo
                                'naturaleza': '',  # Se obtendrá del catálogo
                                'tipo_estado': 'RES',
                                'debe': str(row[2]) if len(row) > 2 and row[2] is not None else '0',  # Columna 3 (índice 2)
                                'haber': str(row[3]) if len(row) > 3 and row[3] is not None else '0'  # Columna 4 (índice 3)
                            })
                
                reader = filas
            else:
                # DEPRECATED: CSV está deprecado, solo Excel es soportado
                messages.error(request, 'El formato CSV está deprecado. Por favor usa el formato Excel (.xlsx o .xls)')
                return redirect('catalogo_upload')
            
            for i, row in enumerate(reader, start=2):
                try:
                    codigo = row.get('codigo', '').strip()
                    nombre = row.get('nombre', '').strip()
                    grupo_nombre = row.get('grupo', '').strip()
                    naturaleza = row.get('naturaleza', '').strip()
                    bg_bloque = row.get('bg_bloque', '').strip()
                    er_bloque = row.get('er_bloque', '').strip()
                    ratio_tag = row.get('ratio_tag', '').strip()
                    tipo_estado = row.get('tipo_estado', 'BAL').strip().upper()
                    
                    # Convertir debe y haber de forma segura
                    try:
                        debe_str = str(row.get('debe', '0') or '0').strip()
                        debe = Decimal(debe_str) if debe_str else Decimal('0')
                    except (ValueError, TypeError, Exception):
                        debe = Decimal('0')
                    
                    try:
                        haber_str = str(row.get('haber', '0') or '0').strip()
                        haber = Decimal(haber_str) if haber_str else Decimal('0')
                    except (ValueError, TypeError, Exception):
                        haber = Decimal('0')
                    
                    if not codigo:
                        errores.append(f"Fila {i}: Falta código de cuenta")
                        continue
                    
                    # Buscar cuenta en el catálogo para obtener información faltante
                    cuenta_existente = None
                    if codigo:
                        cuenta_existente = Cuenta.objects.filter(
                            grupo__catalogo=catalogo,
                            codigo=codigo
                        ).select_related('grupo').first()
                    
                    # Si la cuenta existe, usar sus datos para completar información faltante
                    if cuenta_existente:
                        if not grupo_nombre:
                            grupo_nombre = cuenta_existente.grupo.nombre
                        if not naturaleza:
                            naturaleza = cuenta_existente.grupo.naturaleza
                        if not nombre:
                            nombre = cuenta_existente.nombre
                        if not bg_bloque and cuenta_existente.bg_bloque:
                            bg_bloque = cuenta_existente.bg_bloque
                        if not er_bloque and cuenta_existente.er_bloque:
                            er_bloque = cuenta_existente.er_bloque
                        if not ratio_tag and cuenta_existente.ratio_tag:
                            ratio_tag = cuenta_existente.ratio_tag
                        cuenta = cuenta_existente
                    else:
                        # Si la cuenta no existe, solo continuar si es paso 3 (estados financieros)
                        # En paso 3, si no existe la cuenta, simplemente la ignoramos
                        if paso == '3':
                            # En estados financieros, si la cuenta no existe, la ignoramos
                            # El subtotal se mostrará en 0
                            continue
                        
                        # En paso 2 (catálogo), necesitamos crear la cuenta
                        # Validar que tengamos los datos mínimos
                        if not nombre or not grupo_nombre:
                            errores.append(f"Fila {i}: Faltan campos obligatorios (nombre o grupo) para crear cuenta {codigo}")
                            continue
                        
                        # Crear o obtener grupo
                        grupo, _ = GrupoCuenta.objects.get_or_create(
                            catalogo=catalogo,
                            nombre=grupo_nombre,
                            defaults={'naturaleza': naturaleza}
                        )
                        
                        # Actualizar naturaleza si no tenía
                        if not grupo.naturaleza and naturaleza:
                            grupo.naturaleza = naturaleza
                            grupo.save()
                        
                        # Crear cuenta
                        cuenta, created = Cuenta.objects.get_or_create(
                            grupo=grupo,
                            codigo=codigo,
                            defaults={
                                'nombre': nombre, 
                                'aparece_en_balance': True,
                                'bg_bloque': bg_bloque if bg_bloque else None,
                                'er_bloque': er_bloque if er_bloque else None,
                                'ratio_tag': ratio_tag if ratio_tag else None
                            }
                        )
                        
                        # Actualizar nombre, bloques y ratio_tag si cambiaron
                        actualizar = False
                        if cuenta.nombre != nombre:
                            cuenta.nombre = nombre
                            actualizar = True
                        if bg_bloque and cuenta.bg_bloque != bg_bloque:
                            cuenta.bg_bloque = bg_bloque
                            actualizar = True
                        if er_bloque and cuenta.er_bloque != er_bloque:
                            cuenta.er_bloque = er_bloque
                            actualizar = True
                        if ratio_tag and cuenta.ratio_tag != ratio_tag:
                            cuenta.ratio_tag = ratio_tag
                            actualizar = True
                        if actualizar:
                            cuenta.save()
                    
                    # En paso 2 solo se crean cuentas, no balances
                    # En paso 3 se crean balances con debe/haber
                    if paso == '3':
                        # Si hay debe/haber, crear balance y detalle
                        # Solo procesar si la cuenta existe (si no existe, se ignoró arriba)
                        # Procesar incluso si debe y haber son 0, para tener el registro completo
                        if cuenta:
                            # Crear o obtener balance según tipo
                            if tipo_estado not in balances_por_tipo:
                                balance, _ = Balance.objects.get_or_create(
                                    empresa=empresa,
                                    periodo=periodo,
                                    tipo_balance=tipo_estado
                                )
                                balances_por_tipo[tipo_estado] = balance
                            else:
                                balance = balances_por_tipo[tipo_estado]
                            
                            # Crear o actualizar detalle de balance
                            detalle, created = BalanceDetalle.objects.get_or_create(
                                balance=balance,
                                cuenta=cuenta,
                                defaults={'debe': debe, 'haber': haber}
                            )
                            if not created:
                                detalle.debe = debe
                                detalle.haber = haber
                                detalle.save()
                    
                    creados += 1
                    
                except Exception as e:
                    errores.append(f"Fila {i}: {str(e)}")
            
            # Recalcular saldos solo en paso 3 (cuando se crean balances)
            if paso == '3':
                for balance in balances_por_tipo.values():
                    recalcular_saldos_detalle(balance)
                
                # Calcular ratios automáticamente después de cargar estados financieros
                if periodo:
                    try:
                        # Calcular ratios para ambos tipos de estado
                        ratios_bal = calcular_y_guardar_ratios(empresa, periodo, tipo_estado='BAL')
                        ratios_res = calcular_y_guardar_ratios(empresa, periodo, tipo_estado='RES')
                        # Contar ratios calculados con valores
                        num_ratios = len([r for r in ratios_bal + ratios_res if r['valor'] is not None])
                        if num_ratios > 0:
                            messages.info(request, f"Ratios calculados automáticamente: {num_ratios} ratios procesados para el período {periodo.anio}.")
                    except Exception as e:
                        # No bloquear si falla el cálculo de ratios, solo mostrar advertencia
                        messages.warning(request, f"Estados financieros cargados correctamente, pero hubo un error al calcular ratios: {str(e)}")
            
            if paso == '2':
                # Paso 2: Solo catálogo (sin debe/haber)
                # Verificar si es la primera vez que se carga el catálogo (no hay mapeos previos)
                from stela.models.finanzas import MapeoCuentaLinea
                mapeos_previos = MapeoCuentaLinea.objects.filter(
                    cuenta__grupo__catalogo=catalogo
                ).count()
                
                # Mapear automáticamente cuentas a líneas de estado basándose en ratio_tag
                total_mapeadas = 0
                try:
                    from stela.services.mapeo_automatico import mapear_cuentas_por_bloques
                    resumen_mapeo = mapear_cuentas_por_bloques(catalogo)
                    total_mapeadas = sum(resumen_mapeo.values())
                    if errores:
                        messages.warning(request, f"Se procesaron {creados} cuentas, pero hubo {len(errores)} errores. {total_mapeadas} cuentas mapeadas automáticamente a líneas de estado.")
                    else:
                        messages.success(request, f"Catálogo cargado correctamente. {creados} cuentas procesadas y {total_mapeadas} cuentas mapeadas automáticamente a líneas de estado.")
                except ValueError as e:
                    # Si faltan líneas de estado, mostrar advertencia pero continuar
                    messages.warning(request, f"Catálogo cargado correctamente. {creados} cuentas procesadas. Advertencia: {str(e)}")
                except Exception as e:
                    # Si hay otro error en el mapeo, registrar pero no bloquear
                    messages.warning(request, f"Catálogo cargado correctamente. {creados} cuentas procesadas. Error en mapeo automático: {str(e)}")
                
                # Si es la primera vez (no había mapeos previos) y se crearon mapeos, mostrar paso 2.5
                # Si ya había mapeos o no se crearon nuevos, ir directo a estados financieros
                if mapeos_previos == 0 and total_mapeadas > 0:
                    # Primera vez con mapeos creados: ir al paso de mapeo (paso 2.5)
                    return redirect(f"{reverse('catalogo_upload')}?paso=2.5&empresa={empresa_nit}&catalogo_id={catalogo.id_catalogo}")
                else:
                    # Ya tiene mapeos o no se crearon: ir directo a estados financieros
                    return redirect(f"{reverse('catalogo_upload')}?paso=3&empresa={empresa_nit}&catalogo_id={catalogo.id_catalogo}")
            else:
                # Paso 3: Estados financieros (con debe/haber)
                if errores:
                    messages.warning(request, f"Se procesaron {creados} filas, pero hubo {len(errores)} errores.")
                else:
                    messages.success(request, f"Estados financieros cargados correctamente. {creados} registros procesados.")
                # Redirigir a los detalles de la empresa en lugar del mapeo
                return redirect('empresa_detalles', empresa_nit=empresa.nit)
                
        except Exception as e:
            messages.error(request, f"Error al procesar el archivo: {str(e)}")
    
    # Contexto para el template
    # se cambio por el nuevo campo many to many
    empresas = Empresa.objects.filter(usuario=request.user).distinct().order_by('razon_social')
    catalogo_id = request.GET.get('catalogo_id', '')
    
    # Si hay empresa en parámetro, verificar si ya tiene catálogo
    empresa_seleccionada = None
    if empresa_nit_param:
        try:
            # se cambio por el nuevo campo many to many
            empresa_seleccionada = get_object_or_404(Empresa.objects.filter(usuario=request.user), nit=empresa_nit_param)
            catalogo_existente = Catalogo.objects.filter(empresa=empresa_seleccionada).first()
            if catalogo_existente:
                # Si ya tiene catálogo y estamos en paso 1, saltar al paso 3
                # Si estamos en paso 2, permitir actualizar el catálogo
                if paso == '1':
                    paso = '3'
                    catalogo_id = str(catalogo_existente.id_catalogo)
                    messages.info(request, f'Esta empresa ya tiene un catálogo cargado. Puedes cargar estados financieros para nuevos períodos.')
        except Empresa.DoesNotExist:
            pass
    
    # Si hay catálogo_id, obtener el catálogo para el paso 3 o 2.5
    catalogo = None
    mapeos_existentes = 0
    # Usar catalogo_id_param si catalogo_id está vacío
    catalogo_id_final = catalogo_id or catalogo_id_param
    if catalogo_id_final:
        try:
            catalogo = Catalogo.objects.get(pk=catalogo_id_final, empresa__usuario=request.user)
            # Verificar si ya hay mapeos (para saber si mostrar paso 2.5)
            from stela.models.finanzas import MapeoCuentaLinea
            mapeos_existentes = MapeoCuentaLinea.objects.filter(
                cuenta__grupo__catalogo=catalogo
            ).count()
        except (Catalogo.DoesNotExist, ValueError):
            pass
    
    context = {
        'empresas_usuario': empresas,
        'paso': paso,
        'empresa_nit': empresa_nit_param or request.GET.get('empresa', ''),
        'catalogo_id': catalogo_id_final,
        'catalogo': catalogo,
        'tiene_catalogo': catalogo is not None,
        'tiene_mapeos': mapeos_existentes > 0,
        'num_mapeos': mapeos_existentes
    }
    return render(request, 'stela/catalogo/upload.html', context)


@access_required('009', stay_on_page=True)
def catalogo_create_manual(request):
    """
    Vista para crear catálogo manualmente.
    """
    if request.method == 'POST':
        form = CatalogoManualForm(request.POST, user=request.user)
        if form.is_valid():
            empresa = form.cleaned_data['empresa']
            
            # Crear catálogo (único por empresa, sin año)
            catalogo, created = Catalogo.objects.get_or_create(
                empresa=empresa
            )
            
            if created:
                messages.success(request, f'Catálogo creado para {empresa.razon_social}')
            else:
                messages.info(request, f'El catálogo para {empresa.razon_social} ya existe')
            
            return redirect('dashboard')
    else:
        form = CatalogoManualForm(user=request.user)
    
    context = {
        'form': form,
        'titulo': 'Crear Catálogo Manualmente'
    }
    return render(request, 'stela/catalogo/create_manual.html', context)


@access_required('010', stay_on_page=True)
def catalogo_mapeo_cuentas(request, catalogo_id):
    """
    Vista para mapear cuentas a líneas de estado (para ratios).
    Pestaña oculta, no aparece en el menú principal.
    
    Permite mapeo manual mediante formulario y re-mapeo automático basado en bloques.
    """
    catalogo = get_object_or_404(Catalogo, pk=catalogo_id, empresa__usuario=request.user)
    
    # Si se solicita re-mapeo automático (parámetro GET)
    if request.GET.get('auto_mapear') == '1':
        try:
            from stela.services.mapeo_automatico import mapear_cuentas_por_bloques
            resumen_mapeo = mapear_cuentas_por_bloques(catalogo)
            total_mapeadas = sum(resumen_mapeo.values())
            messages.success(request, f'Mapeo automático completado. {total_mapeadas} cuentas mapeadas a líneas de estado.')
        except ValueError as e:
            messages.error(request, f'Error en mapeo automático: {str(e)}')
        except Exception as e:
            messages.error(request, f'Error inesperado en mapeo automático: {str(e)}')
        # Redirigir para recargar el formulario con los nuevos mapeos
        return redirect(f"{reverse('catalogo_mapeo', args=[catalogo_id])}")
    
    if request.method == 'POST':
        form = MapeoCuentaForm(request.POST, catalogo=catalogo)
        if form.is_valid():
            from stela.models.finanzas import LineaEstado, MapeoCuentaLinea
            
            # Procesar cada campo del formulario (ahora puede tener múltiples cuentas)
            for field_name, cuentas_seleccionadas in form.cleaned_data.items():
                if field_name.startswith('linea_') and cuentas_seleccionadas:
                    clave_linea = field_name.replace('linea_', '')
                    try:
                        linea = LineaEstado.objects.get(clave=clave_linea)
                        
                        # Eliminar TODOS los mapeos existentes para esta línea en este catálogo
                        MapeoCuentaLinea.objects.filter(
                            linea=linea,
                            cuenta__grupo__catalogo=catalogo
                        ).delete()
                        
                        # Crear nuevos mapeos para cada cuenta seleccionada
                        cuentas_mapeadas = 0
                        for cuenta in cuentas_seleccionadas:
                            # Para UTILIDAD_NETA, el signo se determina por la naturaleza de la cuenta
                            # pero se aplicará correctamente en estado_dict
                            signo = 1  # Por defecto positivo
                            if clave_linea == 'UTILIDAD_NETA':
                                # El signo se ajustará en estado_dict según naturaleza y bloque
                                # Aquí solo guardamos 1, el cálculo real se hace en estado_dict
                                signo = 1
                            
                            MapeoCuentaLinea.objects.create(
                                cuenta=cuenta,
                                linea=linea,
                                signo=signo
                            )
                            cuentas_mapeadas += 1
                        
                        if cuentas_mapeadas > 0:
                            messages.info(request, f'{linea.nombre}: {cuentas_mapeadas} cuenta(s) mapeada(s)')
                    except LineaEstado.DoesNotExist:
                        pass
            
            messages.success(request, 'Mapeo de cuentas guardado correctamente.')
            return redirect('dashboard')
    else:
        # Si es la primera vez que se abre esta pantalla y no hay mapeos,
        # ejecutar mapeo automático silenciosamente
        from stela.models.finanzas import MapeoCuentaLinea
        mapeos_existentes = MapeoCuentaLinea.objects.filter(
            cuenta__grupo__catalogo=catalogo
        ).count()
        
        if mapeos_existentes == 0:
            # No hay mapeos, ejecutar automáticamente
            try:
                from stela.services.mapeo_automatico import mapear_cuentas_por_bloques
                resumen_mapeo = mapear_cuentas_por_bloques(catalogo)
                total_mapeadas = sum(resumen_mapeo.values())
                if total_mapeadas > 0:
                    messages.info(request, f'Mapeo automático ejecutado: {total_mapeadas} cuentas mapeadas según su ratio_tag.')
            except Exception as e:
                # Si falla el mapeo automático, continuar sin error
                pass
        
        form = MapeoCuentaForm(catalogo=catalogo)
    
    # Ejecutar mapeo automático si se solicita
    if request.GET.get('auto_mapear') == '1':
        try:
            from stela.services.mapeo_automatico import mapear_cuentas_por_bloques
            resumen_mapeo = mapear_cuentas_por_bloques(catalogo)
            total_mapeadas = sum(resumen_mapeo.values())
            if total_mapeadas > 0:
                messages.success(request, f'Mapeo automático ejecutado: {total_mapeadas} cuentas mapeadas según su ratio_tag, bg_bloque y er_bloque.')
                # Recargar el formulario con los nuevos mapeos
                form = MapeoCuentaForm(catalogo=catalogo)
        except Exception as e:
            messages.error(request, f'Error al ejecutar mapeo automático: {str(e)}')
    
    # Preparar información adicional de cuentas para el template
    # Crear un diccionario con información de grupo y naturaleza para cada cuenta
    cuentas_info = {}
    if catalogo:
        from stela.models.catalogo import Cuenta
        cuentas = Cuenta.objects.filter(grupo__catalogo=catalogo).select_related('grupo')
        for cuenta in cuentas:
            # Usar str(id_cuenta) como clave porque los valores del formulario son strings
            cuentas_info[str(cuenta.id_cuenta)] = {
                'grupo': cuenta.grupo.nombre,
                'naturaleza': cuenta.grupo.get_naturaleza_display() if hasattr(cuenta.grupo, 'get_naturaleza_display') else cuenta.grupo.naturaleza
            }
    
    context = {
        'form': form,
        'catalogo': catalogo,
        'titulo': 'Mapeo de Cuentas para Ratios',
        'cuentas_info': cuentas_info
    }
    return render(request, 'stela/catalogo/mapeo_cuentas.html', context)


# ========== VISTAS PARA DESCARGAR PLANTILLAS ==========

@login_required
def descargar_plantilla_catalogo_csv(request):
    """
    DEPRECATED: Esta función está deprecada. Usa descargar_plantilla_catalogo_excel en su lugar.
    Descarga plantilla CSV de catálogo (solo para compatibilidad hacia atrás).
    """
    # Deprecado: redirigir a Excel o mostrar mensaje
    messages.warning(request, 'El formato CSV está deprecado. Por favor usa el formato Excel.')
    return redirect('descargar_plantilla_catalogo_excel')


@login_required
def descargar_plantilla_catalogo_excel(request):
    """Descarga plantilla Excel de catálogo"""
    output = generar_plantilla_catalogo_excel()
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="plantilla_catalogo.xlsx"'
    return response


@login_required
def descargar_plantilla_estados_csv(request):
    """
    DEPRECATED: Esta función está deprecada. Usa descargar_plantilla_estados_excel en su lugar.
    Descarga plantilla CSV de estados financieros (solo para compatibilidad hacia atrás).
    """
    catalogo_id = request.GET.get('catalogo_id', '')
    if catalogo_id:
        # Deprecado: redirigir a Excel
        messages.warning(request, 'El formato CSV está deprecado. Por favor usa el formato Excel.')
        return redirect('descargar_plantilla_estados_excel', catalogo_id=catalogo_id)
    else:
        messages.warning(request, 'El formato CSV está deprecado. Por favor usa el formato Excel.')
        return redirect('dashboard')


@login_required
def descargar_plantilla_estados_excel(request, catalogo_id):
    """Descarga plantilla Excel de estados financieros basada en catálogo"""
    catalogo = get_object_or_404(Catalogo, pk=catalogo_id, empresa__usuario=request.user)
    output = generar_plantilla_estados_excel(catalogo)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="plantilla_estados_{catalogo.empresa.nit}.xlsx"'
    return response


@login_required
def get_ratios_api(request):
    """
    API que devuelve la lista de TODAS las definiciones de ratios
    (Esto es global, así que está bien como estaba).
    """
    try:
        ratios = RatioDef.objects.all().values(
            'clave', 'nombre', 'formula'
        )
        return JsonResponse(list(ratios), safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def get_cuentas_api(request):
    """
    API que devuelve la lista de cuentas del catálogo
    DE LA EMPRESA ACTIVA (guardada en sesión).
    """
    try:
        # --- Obtener empresa activa de la sesión ---
        active_nit = request.session.get('active_company_nit')
        if not active_nit:
            return JsonResponse({'error': 'No hay empresa activa seleccionada'}, status=404)

        empresa = get_object_or_404(Empresa, nit=active_nit, usuario=request.user)
        # --- Fin de la obtención ---

        catalogo = Catalogo.objects.filter(empresa=empresa).first()
        if not catalogo:
            return JsonResponse([], safe=False)

        # --- ¡CAMBIO AQUÍ! ---
        # Renombramos 'id_cuenta' (de la BD) a 'id' (para el JS)
        cuentas = Cuenta.objects.filter(grupo__catalogo=catalogo).annotate(
            id=F('id_cuenta')
        ).values(
            'id', 'codigo', 'nombre' # Esto ahora funciona gracias al annotate
        )
        # --- FIN DEL CAMBIO ---

        return JsonResponse(list(cuentas), safe=False)

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def get_chart_data_api(request):
    """
    API que devuelve los datos (labels y datasets) para un conjunto
    de ratios o cuentas DE LA EMPRESA ACTIVA.
    """
    try:
        data_type = request.GET.get('type')
        item_ids = request.GET.getlist('ids')

        if not data_type or not item_ids:
            return JsonResponse({'error': 'Faltan parámetros type o ids'}, status=400)

        # --- Obtener empresa activa de la sesión ---
        active_nit = request.session.get('active_company_nit')
        if not active_nit:
            return JsonResponse({'error': 'No hay empresa activa seleccionada'}, status=404)

        # Asumimos que la relación de usuario a empresa es 'usuario' (ManyToMany)
        empresa = get_object_or_404(Empresa.objects.filter(usuario=request.user), nit=active_nit)
        # --- Fin de la obtención ---

        periodos = Periodo.objects.filter(empresa=empresa).order_by('anio', 'mes')
        if not periodos.exists():
            return JsonResponse({'labels': [], 'datasets': []})  # No hay datos para graficar

        labels = [f"{p.anio}-{p.mes:02d}" if p.mes else str(p.anio) for p in periodos]
        datasets = []

        if data_type == 'ratios':
            # --- Lógica de Ratios ---
            for ratio_clave in item_ids:
                try:
                    ratio_def = RatioDef.objects.get(clave=ratio_clave)
                    valores = []
                    for p in periodos:
                        # Buscamos el valor del ratio para ese período
                        r_res = ResultadoRatio.objects.filter(
                            empresa=empresa,
                            periodo=p,
                            ratio=ratio_def
                        ).first()
                        # Añadimos el valor o 0 si no existe
                        valores.append(float(r_res.valor) if r_res and r_res.valor is not None else 0)

                    datasets.append({
                        'label': ratio_def.nombre,
                        'data': valores,
                    })
                except RatioDef.DoesNotExist:
                    pass  # Ignora si la clave del ratio es incorrecta

        elif data_type == 'cuentas':
            # --- Lógica de Cuentas (Corregida) ---
            for cuenta_id in item_ids:
                try:
                    # ¡CORREGIDO! Buscamos por el nombre real del campo en la BD
                    cuenta = Cuenta.objects.get(id_cuenta=cuenta_id)

                    # Validar que la cuenta pertenezca a la empresa activa
                    if cuenta.grupo.catalogo.empresa != empresa:
                        continue  # Esta cuenta no es de esta empresa

                    valores = []
                    for p in periodos:
                        # Buscamos el saldo de esa cuenta para ese período
                        bd = BalanceDetalle.objects.filter(
                            balance__empresa=empresa,
                            balance__periodo=p,
                            cuenta=cuenta
                        ).first()
                        # Añadimos el saldo o 0 si no existe
                        valores.append(float(bd.saldo) if bd and bd.saldo is not None else 0)

                    datasets.append({
                        'label': f"{cuenta.codigo} - {cuenta.nombre}",
                        'data': valores,
                    })
                except Cuenta.DoesNotExist:
                    pass  # Ignora si el ID de cuenta es incorrecto

        return JsonResponse({'labels': labels, 'datasets': datasets})

    except Exception as e:
        # Devuelve el error de Django (como el de 'id' vs 'id_cuenta') al JS
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def set_active_company(request, empresa_nit):
    """
    Guarda el NIT de la empresa seleccionada en la sesión del usuario
    y lo redirige.
    """
    try:
        # 1. Verificar que la empresa existe y que el usuario tiene acceso a ella
        empresa = get_object_or_404(
            Empresa.objects.filter(usuario=request.user),
            nit=empresa_nit
        )

        # 2. Guardar el NIT en la sesión
        request.session['active_company_nit'] = empresa.nit
        messages.success(request, f'Empresa activa cambiada a: {empresa.razon_social}')

    except Exception as e:
        messages.error(request, 'No se pudo seleccionar la empresa.')

    # 3. Redirigir. 'HTTP_REFERER' lo devuelve a la página donde estaba.
    #    Usamos el dashboard como fallback.
    referer = request.META.get('HTTP_REFERER')
    if referer:
        return HttpResponseRedirect(referer)
    return redirect('dashboard')


@login_required
def get_periodos_api(request):
    """
    API que devuelve los períodos disponibles para una empresa según el tipo de estado.
    Solo devuelve períodos que tienen balances del tipo especificado.
    """
    try:
        empresa_id = request.GET.get('empresa_id')
        tipo_estado = request.GET.get('tipo_estado', 'RES')
        
        if not empresa_id:
            return JsonResponse({'error': 'Falta parámetro empresa_id'}, status=400)
        
        # Verificar que la empresa pertenece al usuario (nit es la primary key)
        empresa = get_object_or_404(
            Empresa.objects.filter(usuario=request.user),
            nit=empresa_id
        )
        
        # Obtener períodos que tienen balances del tipo especificado
        periodos = Periodo.objects.filter(
            empresa=empresa,
            balance__tipo_balance=tipo_estado
        ).distinct().order_by('-anio', '-mes')
        
        periodos_data = []
        for periodo in periodos:
            periodo_str = f"{periodo.anio}"
            if periodo.mes:
                periodo_str += f"-{periodo.mes:02d}"
            else:
                periodo_str += " (Anual)"
            
            periodos_data.append({
                'id': periodo.id_periodo,
                'label': periodo_str,
                'anio': periodo.anio,
                'mes': periodo.mes,
            })
        
        return JsonResponse({'periodos': periodos_data}, safe=False)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def ratios_series_json(request):
    """
    API que devuelve datos de ratios para múltiples períodos de una empresa.
    Usado para generar gráficas de evolución temporal.
    """
    try:
        empresa_nit = request.GET.get('empresa')
        ratio_claves = request.GET.get('claves', '').split(',')
        ratio_claves = [c.strip() for c in ratio_claves if c.strip()]
        
        if not empresa_nit:
            return JsonResponse({'error': 'Falta parámetro empresa'}, status=400)
        
        # Validar acceso a la empresa
        empresa = get_object_or_404(
            Empresa.objects.filter(usuario=request.user),
            nit=empresa_nit
        )
        
        # Obtener todos los períodos de la empresa ordenados
        periodos = Periodo.objects.filter(empresa=empresa).order_by('anio', 'mes')
        
        if not periodos.exists():
            return JsonResponse({'anios': [], 'series': {}})
        
        # Crear labels para los períodos
        anios = []
        for p in periodos:
            if p.mes:
                anios.append(f"{p.anio}-{p.mes:02d}")
            else:
                anios.append(str(p.anio))
        
        # Si no se especifican ratios, usar los 5 predefinidos
        if not ratio_claves:
            ratio_claves = [
                'LIQUIDEZ_CORRIENTE',
                'ENDEUDAMIENTO',
                'MARGEN_NETO',
                'ROA',
                'ROE'
            ]
        
        # Obtener datos de cada ratio
        series = {}
        for ratio_clave in ratio_claves:
            try:
                ratio_def = RatioDef.objects.get(clave=ratio_clave)
                valores = []
                for p in periodos:
                    r_res = ResultadoRatio.objects.filter(
                        empresa=empresa,
                        periodo=p,
                        ratio=ratio_def
                    ).first()
                    
                    if r_res and r_res.valor is not None:
                        # Usar valor guardado (más rápido)
                        valores.append(float(r_res.valor))
                    else:
                        # Calcular en tiempo real si no está guardado (respaldo)
                        try:
                            ratios_calc = calcular_y_guardar_ratios(empresa, p, tipo_estado='RES')
                            ratio_encontrado = next((r for r in ratios_calc if r['clave'] == ratio_clave), None)
                            if ratio_encontrado and ratio_encontrado['valor'] is not None:
                                valores.append(float(ratio_encontrado['valor']))
                            else:
                                valores.append(None)
                        except Exception:
                            valores.append(None)
                
                series[ratio_clave] = {
                    'nombre': ratio_def.nombre,
                    'valores': valores
                }
            except RatioDef.DoesNotExist:
                continue
        
        return JsonResponse({
            'anios': anios,
            'series': series
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@access_required('040', stay_on_page=True)
def ratios_graficas_modal(request):
    """
    Vista que valida y devuelve datos para el modal de gráficas de ratios.
    Valida que la empresa tenga al menos 3 períodos con ratios calculados.
    """
    try:
        empresa_nit = request.GET.get('empresa')
        if not empresa_nit:
            return JsonResponse({'error': 'Falta parámetro empresa'}, status=400)
        
        empresa = get_object_or_404(
            Empresa.objects.filter(usuario=request.user),
            nit=empresa_nit
        )
        
        # Contar períodos con ratios calculados
        periodos_con_ratios = Periodo.objects.filter(
            empresa=empresa,
            resultadoratio__isnull=False
        ).distinct().count()
        
        if periodos_con_ratios < 3:
            return JsonResponse({
                'error': f'Se requieren al menos 3 períodos con ratios calculados. Actualmente hay {periodos_con_ratios} períodos.'
            }, status=400)
        
        # Obtener los 5 ratios predefinidos
        ratios_claves = [
            'LIQUIDEZ_CORRIENTE',
            'ENDEUDAMIENTO',
            'MARGEN_NETO',
            'ROA',
            'ROE'
        ]
        
        ratios_info = []
        for clave in ratios_claves:
            try:
                ratio_def = RatioDef.objects.get(clave=clave)
                ratios_info.append({
                    'clave': clave,
                    'nombre': ratio_def.nombre
                })
            except RatioDef.DoesNotExist:
                continue
        
        return JsonResponse({
            'empresa_nit': empresa_nit,
            'empresa_nombre': empresa.razon_social,
            'ratios': ratios_info,
            'periodos_count': periodos_con_ratios
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
